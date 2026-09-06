"""
End-to-end API flow tests for the LMS backend.

Run with:
    python manage.py test test_api_flows
"""
import io
from decimal import Decimal
from unittest.mock import patch

from django.core import mail
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
from rest_framework.test import APITestCase
from rest_framework_simplejwt.tokens import RefreshToken

from accounts.models import Organization, User
from accounts.services import UserProvisioningError, provision_demo_user
from assessments.models import CategorizeItem, CategoryBucket, Choice, HotspotRegion, Question, Quiz, QuizAttempt, WordBankToken
from assignments.models import Assignment, AssignmentSubmission
from audit.models import AuditLog
from certificates.models import Certificate, CertificateTemplate
from certificates.services import (
    MIN_AUTO_SHRINK_FONT_SIZE,
    CertificateIssuanceError,
    _fit_font,
    _resolve_template,
    certificate_ineligibility_reason,
    generate_certificate,
)
from courses.models import (
    Course,
    CourseAccess,
    DemoLessonAccess,
    Element,
    Enrollment,
    Lesson,
    LessonProgress,
    Module,
    Slide,
    SlideProgress,
)
from courses.video_streaming import build_video_stream_token
from gamification.models import Badge, LeaderboardEntry, UserBadge
from gamification.services import (
    COURSE_COMPLETION_POINTS,
    LEVEL_ASSESSMENT_PASS_POINTS,
    award_badges_for_level_assessment_attempt,
    recalculate_leaderboard_entry,
)
from levelassessments.models import (
    AssessmentLevel,
    LevelAssessmentAnswer,
    LevelAssessmentAttempt,
    LevelChoice,
    LevelQuestion,
    QuestionSet,
)
from levelassessments.services import LevelAssessmentError, start_level_assessment_attempt
from narration.models import SlideNarration
from scenarios.models import ScenarioAttempt, ScenarioChoice, ScenarioNode


def make_test_certificate_template(**overrides):
    """Builds a minimal, valid CertificateTemplate for tests that don't rely on the seeded platform default."""
    image_buffer = io.BytesIO()
    Image.new('RGB', (400, 300), color='white').save(image_buffer, format='PNG')
    defaults = dict(
        name='Test Template',
        background_image=SimpleUploadedFile('bg.png', image_buffer.getvalue(), content_type='image/png'),
        is_default=False,
    )
    defaults.update(overrides)
    return CertificateTemplate.objects.create(**defaults)


LEVEL_QUESTION_TEMPLATE_HEADER = [
    'Question Set', 'Question Text', 'Question Type', 'Option A', 'Option B', 'Option C', 'Option D', 'Option E',
    'Correct Answer(s)', 'Marks', 'Explanation', 'Feedback if Correct', 'Feedback if Incorrect',
]


def make_question_template_upload(rows_by_sheet, filename='questions.xlsx'):
    """
    Builds an in-memory .xlsx upload matching the Level Assessment Question
    Template's column structure. `rows_by_sheet` is {sheet_name: [row_tuple, ...]}
    where each row_tuple is 13 values in LEVEL_QUESTION_TEMPLATE_HEADER order.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in rows_by_sheet.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(LEVEL_QUESTION_TEMPLATE_HEADER)
        for row in rows:
            sheet.append(list(row))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(
        filename, buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


class BaseAPITestCase(APITestCase):
    def setUp(self):
        self.org = Organization.objects.create(name='Acme Bank', slug='acme-bank')
        self.other_org = Organization.objects.create(name='Other Bank', slug='other-bank')

        self.learner = User.objects.create_user(
            email='learner@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org,
            first_name='Lana', last_name='Learner',
        )
        self.other_org_learner = User.objects.create_user(
            email='other-learner@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.other_org,
        )
        self.org_admin = User.objects.create_user(
            email='orgadmin@example.com', password='pass12345',
            role=User.Role.ORG_ADMIN, organization=self.org,
        )
        self.instructor = User.objects.create_user(
            email='instructor@example.com', password='pass12345',
            role=User.Role.INSTRUCTOR, organization=self.org,
        )
        self.platform_admin = User.objects.create_user(
            email='platformadmin@example.com', password='pass12345',
            role=User.Role.PLATFORM_ADMIN,
        )

        self.published_org_course = Course.objects.create(
            title='Org Onboarding', slug='org-onboarding', organization=self.org,
            content_owner=Course.ContentOwner.ORGANIZATION, is_published=True,
        )
        self.unpublished_org_course = Course.objects.create(
            title='Org Draft', slug='org-draft', organization=self.org,
            content_owner=Course.ContentOwner.ORGANIZATION, is_published=False,
        )
        self.platform_course = Course.objects.create(
            title='Platform Basics', slug='platform-basics',
            content_owner=Course.ContentOwner.PLATFORM, is_published=True,
        )
        self.other_org_course = Course.objects.create(
            title='Other Org Course', slug='other-org-course', organization=self.other_org,
            content_owner=Course.ContentOwner.ORGANIZATION, is_published=True,
        )

        self.module = Module.objects.create(course=self.published_org_course, title='Intro', order=1)
        self.lesson1 = Lesson.objects.create(module=self.module, title='Welcome', order=1, estimated_minutes=5)
        self.lesson2 = Lesson.objects.create(module=self.module, title='Next steps', order=2, estimated_minutes=5)

        self.quiz_slide = Slide.objects.create(
            lesson=self.lesson1, order=99, title='Final Exam', slide_type=Slide.SlideType.QUIZ,
        )
        self.quiz = Quiz.objects.create(slide=self.quiz_slide, title='Final Exam', pass_percentage=50, max_attempts=2)
        self.q1 = Question.objects.create(quiz=self.quiz, question_text='2+2=?', order=1, points=1)
        self.q1_wrong = Choice.objects.create(question=self.q1, choice_text='3', is_correct=False)
        self.q1_right = Choice.objects.create(question=self.q1, choice_text='4', is_correct=True)
        self.q2 = Question.objects.create(quiz=self.quiz, question_text='Sky color?', order=2, points=1)
        self.q2_right = Choice.objects.create(question=self.q2, choice_text='Blue', is_correct=True)
        self.q2_wrong = Choice.objects.create(question=self.q2, choice_text='Green', is_correct=False)

    def auth_as(self, user):
        access = str(RefreshToken.for_user(user).access_token)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access}')


class AuthFlowTests(BaseAPITestCase):
    def test_login_returns_access_and_refresh_tokens(self):
        response = self.client.post('/api/auth/login/', {'email': 'learner@example.com', 'password': 'pass12345'})
        self.assertEqual(response.status_code, 200)
        self.assertIn('access', response.data)
        self.assertIn('refresh', response.data)

    def test_login_rejects_wrong_password(self):
        response = self.client.post('/api/auth/login/', {'email': 'learner@example.com', 'password': 'wrong'})
        self.assertEqual(response.status_code, 401)

    def test_refresh_returns_new_access_token(self):
        login = self.client.post('/api/auth/login/', {'email': 'learner@example.com', 'password': 'pass12345'})
        response = self.client.post('/api/auth/refresh/', {'refresh': login.data['refresh']})
        self.assertEqual(response.status_code, 200)
        self.assertIn('access', response.data)

    def test_me_returns_current_user_profile(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/auth/me/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['email'], 'learner@example.com')
        self.assertEqual(response.data['role'], 'LEARNER')
        self.assertEqual(response.data['organization']['slug'], 'acme-bank')

    def test_me_requires_authentication(self):
        response = self.client.get('/api/auth/me/')
        self.assertEqual(response.status_code, 401)


class CourseVisibilityTests(BaseAPITestCase):
    def test_learner_sees_only_published_own_org_courses_without_a_grant(self):
        self.auth_as(self.learner)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertEqual(slugs, {'org-onboarding'})

    def test_org_admin_sees_unpublished_own_org_courses_without_a_grant(self):
        self.auth_as(self.org_admin)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertEqual(slugs, {'org-onboarding', 'org-draft'})

    def test_platform_course_becomes_visible_after_grant(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.learner)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertEqual(slugs, {'org-onboarding', 'platform-basics'})

    def test_platform_course_invisible_to_ungranted_org(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.other_org_learner)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertNotIn('platform-basics', slugs)

    def test_platform_admin_sees_every_course(self):
        self.auth_as(self.platform_admin)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertEqual(slugs, {'org-onboarding', 'org-draft', 'platform-basics', 'other-org-course'})

    def test_learner_cannot_see_other_org_course(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/courses/other-org-course/')
        self.assertEqual(response.status_code, 404)

    def test_course_retrieve_returns_nested_modules_and_lessons(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/courses/org-onboarding/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['modules']), 1)
        self.assertEqual(response.data['modules'][0]['lessons'][0]['title'], 'Welcome')


class OrganizationAccessHardeningTests(BaseAPITestCase):
    """
    Explicit query/permission-layer coverage for cross-Organization access:
    a user authenticated as a member of one Organization must never be able
    to retrieve, list, or modify a Course (or its sub-resources) that belongs
    to a different Organization, regardless of role.
    """

    def test_learner_direct_retrieve_of_another_orgs_course_is_denied(self):
        self.auth_as(self.learner)  # self.org
        response = self.client.get(f'/api/courses/{self.other_org_course.slug}/')
        self.assertEqual(response.status_code, 404)

    def test_learner_from_the_target_org_can_retrieve_the_same_course(self):
        self.auth_as(self.other_org_learner)  # self.other_org, matches other_org_course
        response = self.client.get(f'/api/courses/{self.other_org_course.slug}/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['slug'], self.other_org_course.slug)

    def test_other_org_learner_cannot_retrieve_this_orgs_course(self):
        self.auth_as(self.other_org_learner)
        response = self.client.get(f'/api/courses/{self.published_org_course.slug}/')
        self.assertEqual(response.status_code, 404)

    def test_org_admin_direct_retrieve_of_another_orgs_course_is_denied(self):
        self.auth_as(self.org_admin)  # self.org
        response = self.client.get(f'/api/courses/{self.other_org_course.slug}/')
        self.assertEqual(response.status_code, 404)

    def test_org_admin_cannot_edit_another_orgs_course(self):
        self.auth_as(self.org_admin)
        response = self.client.patch(f'/api/courses/{self.other_org_course.slug}/', {'title': 'Hacked'})
        self.assertEqual(response.status_code, 404)
        self.other_org_course.refresh_from_db()
        self.assertEqual(self.other_org_course.title, 'Other Org Course')

    def test_other_org_course_excluded_from_list_endpoint(self):
        self.auth_as(self.learner)
        slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
        self.assertNotIn(self.other_org_course.slug, slugs)

    def test_platform_admin_can_still_retrieve_any_orgs_course(self):
        self.auth_as(self.platform_admin)
        response = self.client.get(f'/api/courses/{self.other_org_course.slug}/')
        self.assertEqual(response.status_code, 200)


class CourseCatalogAndDashboardVisibilityTests(BaseAPITestCase):
    """
    Phase 32 audit: every course-listing surface in the product (the course
    catalog page, the learner/admin dashboard's "recent"/"my courses"
    widgets) is built entirely from GET /api/courses/ (CourseViewSet, scoped
    by visible_courses_for_user) plus GET /api/enrollments/ (scoped by
    RoleScopedQuerysetMixin) — there is no separate catalog/search/
    recommendation backend endpoint to audit independently. These tests
    pin that a standard (non-demo) user's course list never contains a
    course outside their own Organization's assignment, including the edge
    case of a stray Enrollment row pointing at an out-of-scope course.
    """

    def test_catalog_excludes_unpublished_other_org_and_ungranted_platform_courses(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.other_org)
        self.auth_as(self.learner)  # self.org — no grant for platform_course
        response = self.client.get('/api/courses/')
        slugs = {c['slug'] for c in response.data}
        self.assertEqual(slugs, {self.published_org_course.slug})
        self.assertNotIn(self.unpublished_org_course.slug, slugs)
        self.assertNotIn(self.other_org_course.slug, slugs)
        self.assertNotIn(self.platform_course.slug, slugs)

    def test_dashboard_course_list_never_exposes_a_course_behind_a_stray_enrollment(self):
        # Simulate a data-integrity edge case (e.g. a prior bug, a revoked
        # CourseAccess grant after enrollment) rather than one reachable
        # through the current API: an Enrollment row pointing at a course
        # outside the enrolled user's org.
        Enrollment.objects.create(user=self.other_org_learner, course=self.published_org_course)

        self.auth_as(self.other_org_learner)
        catalog = self.client.get('/api/courses/')
        self.assertNotIn(self.published_org_course.slug, {c['slug'] for c in catalog.data})

        enrollments = self.client.get('/api/enrollments/')
        self.assertEqual(enrollments.data[0]['course'], self.published_org_course.id)
        self.assertNotIn('title', enrollments.data[0])

    def test_ungranted_platform_course_absent_from_catalog_for_every_organization(self):
        for user in (self.learner, self.other_org_learner):
            self.auth_as(user)
            slugs = {c['slug'] for c in self.client.get('/api/courses/').data}
            self.assertNotIn(self.platform_course.slug, slugs)


class EnrollmentFlowTests(BaseAPITestCase):
    def test_learner_can_enroll_and_list_own_enrollment(self):
        self.auth_as(self.learner)
        create = self.client.post('/api/enrollments/', {'course': self.published_org_course.id})
        self.assertEqual(create.status_code, 201)

        listing = self.client.get('/api/enrollments/')
        self.assertEqual(len(listing.data), 1)
        self.assertEqual(listing.data[0]['course'], self.published_org_course.id)

    def test_duplicate_enrollment_rejected(self):
        self.auth_as(self.learner)
        self.client.post('/api/enrollments/', {'course': self.published_org_course.id})
        response = self.client.post('/api/enrollments/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 400)

    def test_cannot_enroll_in_course_outside_org_access(self):
        self.auth_as(self.learner)
        response = self.client.post('/api/enrollments/', {'course': self.other_org_course.id})
        self.assertEqual(response.status_code, 400)

    def test_update_progress_to_completed_sets_completed_at(self):
        self.auth_as(self.learner)
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        response = self.client.patch(f'/api/enrollments/{enrollment.id}/', {'status': 'COMPLETED'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['status'], 'COMPLETED')
        self.assertIsNotNone(response.data['completed_at'])
        self.assertEqual(response.data['progress_percent'], 100)

    def test_other_learner_cannot_see_or_edit_someone_elses_enrollment(self):
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        self.auth_as(self.other_org_learner)
        self.assertEqual(self.client.get(f'/api/enrollments/{enrollment.id}/').status_code, 404)
        self.assertEqual(self.client.patch(f'/api/enrollments/{enrollment.id}/', {'progress_percent': 50}).status_code, 404)

    def test_org_admin_sees_enrollments_within_their_org_only(self):
        Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        Enrollment.objects.create(user=self.other_org_learner, course=self.other_org_course)
        self.auth_as(self.org_admin)
        response = self.client.get('/api/enrollments/')
        self.assertEqual({row['user'] for row in response.data}, {self.learner.id})

    def test_platform_admin_sees_all_enrollments(self):
        Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        Enrollment.objects.create(user=self.other_org_learner, course=self.other_org_course)
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/enrollments/')
        self.assertEqual(len(response.data), 2)

    def test_list_enrollments_filtered_by_course_query_param(self):
        Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        Enrollment.objects.create(user=self.learner, course=self.platform_course)
        self.auth_as(self.learner)
        response = self.client.get(f'/api/enrollments/?course={self.published_org_course.id}')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['course'], self.published_org_course.id)

    def test_complete_lesson_updates_progress_and_completed_ids(self):
        self.auth_as(self.learner)
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)

        response = self.client.post(
            f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['completed_lesson_ids'], [self.lesson1.id])
        self.assertEqual(response.data['progress_percent'], 50)
        self.assertEqual(response.data['status'], 'IN_PROGRESS')

    def test_completing_all_lessons_marks_enrollment_completed(self):
        self.auth_as(self.learner)
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)

        self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id})
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson2.id})

        self.assertEqual(response.data['progress_percent'], 100)
        self.assertEqual(response.data['status'], 'COMPLETED')
        self.assertIsNotNone(response.data['completed_at'])
        self.assertCountEqual(response.data['completed_lesson_ids'], [self.lesson1.id, self.lesson2.id])

    def test_complete_lesson_is_idempotent(self):
        self.auth_as(self.learner)
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id})
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id})
        self.assertEqual(response.data['completed_lesson_ids'], [self.lesson1.id])
        self.assertEqual(response.data['progress_percent'], 50)

    def test_cannot_complete_lesson_from_a_different_course(self):
        self.auth_as(self.learner)
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        other_module = Module.objects.create(course=self.platform_course, title='Other', order=1)
        other_lesson = Lesson.objects.create(module=other_module, title='Other lesson', order=1)
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': other_lesson.id})
        self.assertEqual(response.status_code, 404)

    def test_other_learner_cannot_complete_lesson_on_someone_elses_enrollment(self):
        enrollment = Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        self.auth_as(self.other_org_learner)
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id})
        self.assertEqual(response.status_code, 404)


class EnrollmentRetakeTests(BaseAPITestCase):
    """"Retake Course" (CourseCompletionModal.tsx) resets an enrollment to a
    fresh state — progress, quiz attempts (so max_attempts starts over), and
    every other attempt type, scoped to just this user+course."""

    def setUp(self):
        super().setUp()
        self.enrollment = Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED, progress_percent=100,
            completed_at=timezone.now(),
        )
        SlideProgress.objects.create(enrollment=self.enrollment, slide=self.quiz_slide, completed_at=timezone.now())
        LessonProgress.objects.create(enrollment=self.enrollment, lesson=self.lesson1, completed_at=timezone.now())
        # Exhaust max_attempts (2) so we can prove a retake actually frees it up.
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=False, score_percent=0)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=2, passed=False, score_percent=0)

        scenario_slide = Slide.objects.create(
            lesson=self.lesson2, order=1, title='Scenario', slide_type=Slide.SlideType.SCENARIO,
        )
        node = ScenarioNode.objects.create(slide=scenario_slide, node_key='start', is_start=True)
        ScenarioAttempt.objects.create(enrollment=self.enrollment, slide=scenario_slide, path_taken=[node.id])

        assignment_slide = Slide.objects.create(
            lesson=self.lesson2, order=2, title='Assignment', slide_type=Slide.SlideType.ASSIGNMENT,
        )
        assignment = Assignment.objects.create(slide=assignment_slide)
        AssignmentSubmission.objects.create(assignment=assignment, user=self.learner, text_response='done')

        # A different course's data for the same learner must survive untouched.
        self.other_enrollment = Enrollment.objects.create(user=self.learner, course=self.platform_course)
        other_module = Module.objects.create(course=self.platform_course, title='M', order=1)
        other_lesson = Lesson.objects.create(module=other_module, title='L', order=1)
        other_quiz_slide = Slide.objects.create(lesson=other_lesson, order=1, title='Q', slide_type=Slide.SlideType.QUIZ)
        self.other_quiz = Quiz.objects.create(slide=other_quiz_slide, title='Q', pass_percentage=50)
        QuizAttempt.objects.create(user=self.learner, quiz=self.other_quiz, attempt_number=1, passed=True, score_percent=100)

    def test_retake_resets_enrollment_and_deletes_all_progress_and_attempts(self):
        self.auth_as(self.learner)
        response = self.client.post(f'/api/enrollments/{self.enrollment.id}/retake/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['status'], 'NOT_STARTED')
        self.assertEqual(response.data['progress_percent'], 0)
        self.assertIsNone(response.data['completed_at'])

        self.assertFalse(SlideProgress.objects.filter(enrollment=self.enrollment).exists())
        self.assertFalse(LessonProgress.objects.filter(enrollment=self.enrollment).exists())
        self.assertFalse(QuizAttempt.objects.filter(user=self.learner, quiz=self.quiz).exists())
        self.assertFalse(ScenarioAttempt.objects.filter(enrollment=self.enrollment).exists())
        self.assertFalse(AssignmentSubmission.objects.filter(user=self.learner).filter(assignment__slide__lesson=self.lesson2).exists())

    def test_retake_lets_learner_exceed_the_old_max_attempts(self):
        self.auth_as(self.learner)
        self.client.post(f'/api/enrollments/{self.enrollment.id}/retake/')

        payload = [{'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                   {'question': self.q2.id, 'selected_choices': [self.q2_right.id]}]
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', {'answers': payload}, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(QuizAttempt.objects.filter(user=self.learner, quiz=self.quiz).count(), 1)

    def test_retake_does_not_touch_a_different_course(self):
        self.auth_as(self.learner)
        self.client.post(f'/api/enrollments/{self.enrollment.id}/retake/')

        self.other_enrollment.refresh_from_db()
        self.assertTrue(QuizAttempt.objects.filter(user=self.learner, quiz=self.other_quiz).exists())

    def test_other_learner_cannot_retake_someone_elses_enrollment(self):
        self.auth_as(self.other_org_learner)
        response = self.client.post(f'/api/enrollments/{self.enrollment.id}/retake/')
        self.assertEqual(response.status_code, 404)
        self.enrollment.refresh_from_db()
        self.assertEqual(self.enrollment.status, 'COMPLETED')


class QuizFlowTests(BaseAPITestCase):
    def test_learner_cannot_see_correct_answers(self):
        self.auth_as(self.learner)
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        self.assertEqual(response.status_code, 200)
        for question in response.data['questions']:
            for choice in question['choices']:
                self.assertNotIn('is_correct', choice)

    def test_org_admin_can_see_correct_answers(self):
        self.auth_as(self.org_admin)
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        for question in response.data['questions']:
            for choice in question['choices']:
                self.assertIn('is_correct', choice)

    def test_submit_scores_attempt_and_marks_passed(self):
        self.auth_as(self.learner)
        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
            ],
        }
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['score_percent'], '100.00')
        self.assertTrue(response.data['passed'])

    def test_submit_with_one_wrong_answer_scores_partial(self):
        self.auth_as(self.learner)
        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_wrong.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
            ],
        }
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(response.data['score_percent'], '50.00')
        self.assertTrue(response.data['passed'])  # pass_percentage is 50

    def test_max_attempts_enforced(self):
        self.auth_as(self.learner)
        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
            ],
        }
        self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        third = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(third.status_code, 400)
        self.assertEqual(QuizAttempt.objects.filter(user=self.learner, quiz=self.quiz).count(), 2)

    def test_submit_rejects_choice_from_a_different_question(self):
        self.auth_as(self.learner)
        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q2_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
            ],
        }
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(response.status_code, 400)


class MultipleAnswerScoringTests(BaseAPITestCase):
    """MULTIPLE_ANSWER questions score all-or-nothing: every correct option
    selected, no incorrect ones — see assessments.views.QuizViewSet.submit."""

    def setUp(self):
        super().setUp()
        self.ma_question = Question.objects.create(
            quiz=self.quiz,
            question_text='Which are primary colors?',
            question_type=Question.QuestionType.MULTIPLE_ANSWER,
            order=3,
            points=1,
            explanation='Red, blue, and yellow are the primary colors.',
            feedback_correct='Nice — you got every primary color.',
            feedback_incorrect='Not quite — check which ones are true primaries.',
        )
        self.red = Choice.objects.create(question=self.ma_question, choice_text='Red', is_correct=True)
        self.blue = Choice.objects.create(question=self.ma_question, choice_text='Blue', is_correct=True)
        self.green = Choice.objects.create(question=self.ma_question, choice_text='Green', is_correct=False)
        self.purple = Choice.objects.create(question=self.ma_question, choice_text='Purple', is_correct=False)

    def _submit(self, selected_choice_ids):
        self.auth_as(self.learner)
        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
                {'question': self.ma_question.id, 'selected_choices': selected_choice_ids},
            ],
        }
        return self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')

    def _ma_answer(self, response):
        return next(a for a in response.data['answers'] if a['question'] == self.ma_question.id)

    def test_selecting_both_correct_options_scores_correct(self):
        response = self._submit([self.red.id, self.blue.id])
        self.assertEqual(response.status_code, 201)
        answer = self._ma_answer(response)
        self.assertTrue(answer['is_correct'])

    def test_selecting_only_one_correct_option_scores_incorrect(self):
        response = self._submit([self.red.id])
        answer = self._ma_answer(response)
        self.assertFalse(answer['is_correct'])

    def test_selecting_a_correct_option_plus_a_distractor_scores_incorrect(self):
        response = self._submit([self.red.id, self.green.id])
        answer = self._ma_answer(response)
        self.assertFalse(answer['is_correct'])

    def test_single_choice_scoring_is_unaffected(self):
        # Same payload shape/logic path as MULTIPLE_ANSWER (exact set
        # equality) — confirms the existing SINGLE_CHOICE behavior wasn't
        # changed by adding MULTIPLE_ANSWER support.
        response = self._submit([self.red.id, self.blue.id])
        q1_answer = next(a for a in response.data['answers'] if a['question'] == self.q1.id)
        self.assertTrue(q1_answer['is_correct'])

    def test_result_reveals_correct_choice_ids_explanation_and_feedback(self):
        response = self._submit([self.red.id, self.blue.id])
        answer = self._ma_answer(response)
        self.assertEqual(set(answer['correct_choice_ids']), {self.red.id, self.blue.id})
        self.assertEqual(answer['explanation'], 'Red, blue, and yellow are the primary colors.')
        self.assertEqual(answer['feedback_correct'], 'Nice — you got every primary color.')
        self.assertEqual(answer['feedback_incorrect'], 'Not quite — check which ones are true primaries.')

    def test_quiz_detail_still_hides_answer_key_from_learner_before_submitting(self):
        # The reveal lives on the QuizAttempt/QuizAnswer response only — the
        # quiz-taking endpoint itself must still strip the answer key.
        self.auth_as(self.learner)
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        ma_data = next(q for q in response.data['questions'] if q['id'] == self.ma_question.id)
        self.assertNotIn('explanation', ma_data)
        self.assertNotIn('feedback_correct', ma_data)
        for choice in ma_data['choices']:
            self.assertNotIn('is_correct', choice)


class ChoiceDisplayOrderTests(BaseAPITestCase):
    """
    Phase 35: SINGLE_CHOICE/MULTIPLE_CHOICE/MULTIPLE_ANSWER/TRUE_FALSE choices
    must not always render in their stored creation order for a learner —
    that lets the correct answer's position (not its content) become a
    learnable pattern. See QuestionSerializer.to_representation.
    """

    def setUp(self):
        super().setUp()
        self.sc_question = Question.objects.create(
            quiz=self.quiz,
            question_text='Which is the capital of France?',
            question_type=Question.QuestionType.SINGLE_CHOICE,
            order=4,
            points=1,
        )
        # The correct choice is deliberately created first — the exact
        # authoring pattern that produced the positional-bias bug, so a
        # regression here would show every fetch's choices[0] as correct.
        self.paris = Choice.objects.create(question=self.sc_question, choice_text='Paris', is_correct=True, order=1)
        self.berlin = Choice.objects.create(question=self.sc_question, choice_text='Berlin', is_correct=False, order=2)
        self.madrid = Choice.objects.create(question=self.sc_question, choice_text='Madrid', is_correct=False, order=3)
        self.rome = Choice.objects.create(question=self.sc_question, choice_text='Rome', is_correct=False, order=4)
        self.lisbon = Choice.objects.create(question=self.sc_question, choice_text='Lisbon', is_correct=False, order=5)
        self.oslo = Choice.objects.create(question=self.sc_question, choice_text='Oslo', is_correct=False, order=6)

    def _fetch_choice_order(self):
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        sc_data = next(q for q in response.data['questions'] if q['id'] == self.sc_question.id)
        return [choice['id'] for choice in sc_data['choices']]

    def test_learner_sees_shuffled_choice_order_across_fetches(self):
        self.auth_as(self.learner)
        orders = [tuple(self._fetch_choice_order()) for _ in range(30)]
        creation_order = (self.paris.id, self.berlin.id, self.madrid.id, self.rome.id, self.lisbon.id, self.oslo.id)

        # Every fetch still contains exactly the same 6 choices...
        for order in orders:
            self.assertEqual(set(order), set(creation_order))
        # ...but with 6 choices shuffled independently 30 times, seeing the
        # exact same order every single time is astronomically unlikely
        # (~1/720 per trial) unless shuffling isn't actually happening.
        self.assertGreater(len(set(orders)), 1)

    def test_correct_choices_position_varies_not_always_first(self):
        self.auth_as(self.learner)
        positions = [self._fetch_choice_order().index(self.paris.id) for _ in range(30)]
        # If the bug were present, every position would be 0 (creation order,
        # correct choice authored first).
        self.assertGreater(len(set(positions)), 1)

    def test_privileged_role_sees_stable_authored_order(self):
        # Instructors/admins are editing, not guessing — shuffling would just
        # make the authoring UI's own choices jump around, so it stays off
        # for privileged roles, same as every other question type here.
        self.auth_as(self.instructor)
        orders = [tuple(self._fetch_choice_order()) for _ in range(5)]
        creation_order = (self.paris.id, self.berlin.id, self.madrid.id, self.rome.id, self.lisbon.id, self.oslo.id)
        self.assertEqual(set(orders), {creation_order})

    def test_grading_is_unaffected_by_display_shuffle(self):
        # Grading is by Choice id set-equality (assessments.views.QuizViewSet
        # .submit), never by the shuffled array position, so this must keep
        # passing regardless of how many times the quiz was re-fetched first.
        self.auth_as(self.learner)
        for _ in range(10):
            self._fetch_choice_order()

        payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
                {'question': self.sc_question.id, 'selected_choices': [self.paris.id]},
            ],
        }
        response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(response.status_code, 201)
        sc_answer = next(a for a in response.data['answers'] if a['question'] == self.sc_question.id)
        self.assertTrue(sc_answer['is_correct'])

        wrong_payload = {
            'answers': [
                {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
                {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
                {'question': self.sc_question.id, 'selected_choices': [self.berlin.id]},
            ],
        }
        wrong_response = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', wrong_payload, format='json')
        wrong_answer = next(a for a in wrong_response.data['answers'] if a['question'] == self.sc_question.id)
        self.assertFalse(wrong_answer['is_correct'])


class CertificateFlowTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        attempt = QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)
        self.certificate = generate_certificate(self.learner, self.published_org_course)
        self.attempt = attempt

    def test_learner_can_list_and_retrieve_own_certificate(self):
        self.auth_as(self.learner)
        listing = self.client.get('/api/certificates/')
        self.assertEqual(len(listing.data), 1)

        retrieve = self.client.get(f'/api/certificates/{self.certificate.id}/')
        self.assertEqual(retrieve.status_code, 200)
        self.assertEqual(retrieve.data['certificate_number'], self.certificate.certificate_number)

    def test_other_learner_cannot_see_certificate(self):
        self.auth_as(self.other_org_learner)
        response = self.client.get(f'/api/certificates/{self.certificate.id}/')
        self.assertEqual(response.status_code, 404)

    def test_download_returns_pdf(self):
        self.auth_as(self.learner)
        response = self.client.get(f'/api/certificates/{self.certificate.id}/download/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_public_verify_endpoint_returns_certificate_details_without_auth(self):
        self.client.credentials()  # no auth header
        response = self.client.get(f'/verify/{self.certificate.verification_token}/')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['valid'])
        self.assertEqual(response.json()['learner_name'], 'Lana Learner')

    def test_public_verify_endpoint_returns_404_for_unknown_token(self):
        self.client.credentials()
        response = self.client.get('/verify/00000000-0000-0000-0000-000000000000/')
        self.assertEqual(response.status_code, 404)
        self.assertFalse(response.json()['valid'])


class CertificateIssueEndpointTests(BaseAPITestCase):
    def test_issue_endpoint_creates_certificate_once_eligible(self):
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)

        self.auth_as(self.learner)
        response = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 200)
        self.assertIn('certificate_number', response.data)
        self.assertTrue(response.data['pdf_file'])

    def test_issue_endpoint_rejects_when_quiz_not_attempted(self):
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        self.auth_as(self.learner)
        response = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 400)
        self.assertIn('has not been attempted yet', response.data['detail'])

    def test_issue_endpoint_is_idempotent(self):
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)

        self.auth_as(self.learner)
        first = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        second = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(first.data['id'], second.data['id'])

    def test_issue_endpoint_rejects_course_outside_requesters_organization(self):
        # Directly construct an eligible-looking Enrollment/QuizAttempt for a
        # learner in a different org than the course, bypassing the normal
        # enroll-time visible_courses_for_user gate, to prove the issue
        # endpoint itself also org-scopes the course lookup as defense in depth.
        Enrollment.objects.create(
            user=self.other_org_learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(
            user=self.other_org_learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100
        )

        self.auth_as(self.other_org_learner)
        response = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 404)


class CourseAverageCertificateEligibilityTests(BaseAPITestCase):
    """
    Phase 34: certificate eligibility is governed by the course-wide AVERAGE
    score across all of the course's quizzes (each quiz's own best attempt),
    not a requirement that every individual quiz independently score above
    its own Quiz.pass_percentage. self.quiz (pass_percentage=50) already
    exists on lesson1 from BaseAPITestCase; a second quiz (pass_percentage=
    70) is added on lesson2 here so the average can diverge from any single
    quiz's individual pass/fail outcome. published_org_course.
    certificate_pass_threshold is the default (70).
    """

    def setUp(self):
        super().setUp()
        self.quiz2_slide = Slide.objects.create(
            lesson=self.lesson2, order=99, title='Second Exam', slide_type=Slide.SlideType.QUIZ,
        )
        self.quiz2 = Quiz.objects.create(slide=self.quiz2_slide, title='Second Exam', pass_percentage=70)
        self.enrollment = Enrollment.objects.create(
            user=self.learner, course=self.published_org_course, status=Enrollment.Status.COMPLETED,
        )

    def test_certificate_issues_when_average_meets_threshold_despite_one_quiz_individually_failed(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=False, score_percent=40)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=True, score_percent=100)
        # Average = (40 + 100) / 2 = 70, meets the 70% threshold — even though
        # self.quiz was individually failed against its own pass_percentage=50.

        self.assertIsNone(certificate_ineligibility_reason(self.learner, self.published_org_course))

        self.auth_as(self.learner)
        response = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Certificate.objects.filter(user=self.learner, course=self.published_org_course).exists())

    def test_certificate_denied_when_average_below_threshold_despite_one_quiz_individually_passed(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=False, score_percent=20)
        # Average = (100 + 20) / 2 = 60, below the 70% threshold — even though
        # self.quiz was individually passed against its own pass_percentage=50.

        reason = certificate_ineligibility_reason(self.learner, self.published_org_course)
        self.assertIsNotNone(reason)
        self.assertIn('60.0%', reason)

        self.auth_as(self.learner)
        response = self.client.post('/api/certificates/issue/', {'course': self.published_org_course.id})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Certificate.objects.filter(user=self.learner, course=self.published_org_course).exists())

    def test_unattempted_quiz_blocks_issuance_even_with_a_high_score_elsewhere(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)
        # self.quiz2 was never attempted at all.

        reason = certificate_ineligibility_reason(self.learner, self.published_org_course)
        self.assertIsNotNone(reason)
        self.assertIn('has not been attempted yet', reason)

    def test_enrollment_serializer_exposes_ineligible_reason_only_once_completed_and_short(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=False, score_percent=20)

        self.auth_as(self.learner)
        response = self.client.get(f'/api/enrollments/{self.enrollment.id}/')
        self.assertIsNotNone(response.data['certificate_ineligible_reason'])
        self.assertIn('60.0%', response.data['certificate_ineligible_reason'])

    def test_enrollment_serializer_reason_is_null_when_average_meets_threshold(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=False, score_percent=40)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=True, score_percent=100)

        self.auth_as(self.learner)
        response = self.client.get(f'/api/enrollments/{self.enrollment.id}/')
        self.assertIsNone(response.data['certificate_ineligible_reason'])

    def test_enrollment_serializer_reason_is_null_while_still_in_progress(self):
        self.enrollment.status = Enrollment.Status.IN_PROGRESS
        self.enrollment.save()

        self.auth_as(self.learner)
        response = self.client.get(f'/api/enrollments/{self.enrollment.id}/')
        self.assertIsNone(response.data['certificate_ineligible_reason'])

    def test_certificate_eligible_when_average_is_exactly_the_70_percent_threshold(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=70)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=True, score_percent=70)
        # Average = (70 + 70) / 2 = 70.0 — the boundary itself. Eligibility is
        # "at or above" the threshold, so this must be eligible, not blocked.

        self.assertIsNone(certificate_ineligibility_reason(self.learner, self.published_org_course))

    def test_certificate_ineligible_when_average_is_just_below_the_70_percent_threshold(self):
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=70)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=False, score_percent=Decimal('69.8'))
        # Average = (70 + 69.8) / 2 = 69.9 — one tenth of a point under the
        # boundary, so this must be ineligible.

        reason = certificate_ineligibility_reason(self.learner, self.published_org_course)
        self.assertIsNotNone(reason)
        self.assertIn('69.9%', reason)


class OrganizationListTests(BaseAPITestCase):
    def test_platform_admin_can_list_organizations(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/organizations/')
        self.assertEqual(response.status_code, 200)
        self.assertIn(self.org.slug, [o['slug'] for o in response.data])

    def test_learner_forbidden_from_organization_list(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/organizations/')
        self.assertEqual(response.status_code, 403)

    def test_cannot_create_organization_with_duplicate_name(self):
        self.auth_as(self.platform_admin)
        response = self.client.post('/api/organizations/', {'name': self.org.name.upper()})
        self.assertEqual(response.status_code, 400)
        self.assertIn('already exists', str(response.data))

    def test_platform_admin_can_delete_empty_organization(self):
        empty_org = Organization.objects.create(name='Empty Org', slug='empty-org')
        self.auth_as(self.platform_admin)
        response = self.client.delete(f'/api/organizations/{empty_org.id}/')
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Organization.objects.filter(id=empty_org.id).exists())

    def test_cannot_delete_organization_with_users(self):
        self.auth_as(self.platform_admin)
        response = self.client.delete(f'/api/organizations/{self.org.id}/')
        self.assertEqual(response.status_code, 400)
        self.assertTrue(Organization.objects.filter(id=self.org.id).exists())

    def test_org_admin_cannot_delete_organization(self):
        empty_org = Organization.objects.create(name='Empty Org 2', slug='empty-org-2')
        self.auth_as(self.org_admin)
        response = self.client.delete(f'/api/organizations/{empty_org.id}/')
        self.assertEqual(response.status_code, 403)


class CourseBuilderTests(BaseAPITestCase):
    def test_learner_cannot_create_course(self):
        self.auth_as(self.learner)
        response = self.client.post('/api/courses/', {'title': 'New', 'slug': 'new-course'})
        self.assertEqual(response.status_code, 403)

    def test_instructor_create_forces_own_organization_and_content_owner(self):
        self.auth_as(self.instructor)
        response = self.client.post('/api/courses/', {
            'title': 'Instructor Course', 'slug': 'instructor-course',
            'organization': self.other_org.id, 'content_owner': 'PLATFORM',
        })
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['organization'], self.org.id)
        self.assertEqual(response.data['content_owner'], 'ORGANIZATION')
        self.assertEqual(Course.objects.get(slug='instructor-course').created_by, self.instructor)

    def test_platform_admin_create_forces_platform_content_owner(self):
        self.auth_as(self.platform_admin)
        response = self.client.post('/api/courses/', {
            'title': 'Platform Managed', 'slug': 'platform-managed',
            'organization': self.other_org.id, 'content_owner': 'ORGANIZATION',
        })
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['organization'], self.other_org.id)
        self.assertEqual(response.data['content_owner'], 'PLATFORM')

    def test_org_admin_cannot_edit_content_owner_on_update(self):
        self.auth_as(self.org_admin)
        response = self.client.patch(
            f'/api/courses/{self.published_org_course.slug}/', {'content_owner': 'PLATFORM'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['content_owner'], 'ORGANIZATION')

    def test_instructor_cannot_edit_other_org_course(self):
        self.auth_as(self.instructor)
        response = self.client.patch(f'/api/courses/{self.other_org_course.slug}/', {'title': 'Hacked'})
        self.assertEqual(response.status_code, 404)

    def test_org_admin_can_edit_draft_course_in_own_org(self):
        self.auth_as(self.org_admin)
        response = self.client.patch(f'/api/courses/{self.unpublished_org_course.slug}/', {'is_published': True})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['is_published'])

    def test_org_admin_cannot_move_course_to_another_org(self):
        self.auth_as(self.org_admin)
        response = self.client.patch(
            f'/api/courses/{self.published_org_course.slug}/', {'organization': self.other_org.id}
        )
        self.assertEqual(response.status_code, 200)
        self.published_org_course.refresh_from_db()
        self.assertEqual(self.published_org_course.organization_id, self.org.id)


class CourseAccessGrantTests(BaseAPITestCase):
    def test_platform_admin_can_grant_and_list_access(self):
        self.auth_as(self.platform_admin)
        grant_response = self.client.post(
            f'/api/courses/{self.platform_course.slug}/access-grants/', {'organization': self.org.id}
        )
        self.assertEqual(grant_response.status_code, 201)

        list_response = self.client.get(f'/api/courses/{self.platform_course.slug}/access-grants/')
        self.assertEqual(len(list_response.data), 1)
        self.assertEqual(list_response.data[0]['organization']['id'], self.org.id)

    def test_granting_is_idempotent(self):
        self.auth_as(self.platform_admin)
        first = self.client.post(f'/api/courses/{self.platform_course.slug}/access-grants/', {'organization': self.org.id})
        second = self.client.post(f'/api/courses/{self.platform_course.slug}/access-grants/', {'organization': self.org.id})
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(CourseAccess.objects.filter(course=self.platform_course, organization=self.org).count(), 1)

    def test_cannot_grant_access_to_an_organization_owned_course(self):
        self.auth_as(self.platform_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/access-grants/', {'organization': self.other_org.id}
        )
        self.assertEqual(response.status_code, 400)

    def test_revoke_removes_the_grant(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.platform_admin)
        response = self.client.delete(
            f'/api/courses/{self.platform_course.slug}/access-grants/revoke/', {'organization': self.org.id},
            format='json',
        )
        self.assertEqual(response.status_code, 204)
        self.assertFalse(CourseAccess.objects.filter(course=self.platform_course, organization=self.org).exists())

    def test_org_admin_cannot_manage_grants_on_platform_course(self):
        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.platform_course.slug}/access-grants/', {'organization': self.org.id}
        )
        # editable_courses_for_user excludes PLATFORM-owned courses for ORG_ADMIN,
        # so the object lookup itself 404s before any content_owner check runs.
        self.assertEqual(response.status_code, 404)

    def test_org_admin_with_granted_access_still_cannot_edit_platform_course(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.org_admin)
        response = self.client.patch(f'/api/courses/{self.platform_course.slug}/', {'title': 'Hacked'})
        self.assertEqual(response.status_code, 404)


class CourseCloneTests(BaseAPITestCase):
    """Covers courses.services.clone_course_for_organization via the /clone/ endpoint,
    exercising one slide of every type so every deep-copy branch runs."""

    def setUp(self):
        super().setUp()
        self.module = Module.objects.create(course=self.platform_course, title='Module 1', order=1)
        self.lesson = Lesson.objects.create(module=self.module, title='Lesson 1', order=1, estimated_minutes=5)

        self.content_slide = Slide.objects.create(lesson=self.lesson, order=1, slide_type=Slide.SlideType.CONTENT)
        Element.objects.create(slide=self.content_slide, order=1, element_type=Element.ElementType.TEXT, rich_text='Hello')

        self.quiz_slide2 = Slide.objects.create(lesson=self.lesson, order=2, slide_type=Slide.SlideType.QUIZ)
        quiz = Quiz.objects.create(slide=self.quiz_slide2, title='Quiz', pass_percentage=60)
        question = Question.objects.create(
            quiz=quiz, question_text='Categorize this', order=1, question_type=Question.QuestionType.CATEGORIZE,
        )
        bucket = CategoryBucket.objects.create(question=question, label='Bucket A', order=1)
        CategorizeItem.objects.create(question=question, item_text='Item 1', correct_bucket=bucket, order=1)
        HotspotRegion.objects.create(question=question, x=1, y=1, width=10, height=10, is_correct=True)
        WordBankToken.objects.create(question=question, text='word', correct_blank_index=1, order=1)
        Choice.objects.create(question=question, choice_text='Option', is_correct=True, order=1)

        self.assignment_slide = Slide.objects.create(lesson=self.lesson, order=3, slide_type=Slide.SlideType.ASSIGNMENT)
        Assignment.objects.create(slide=self.assignment_slide, instructions='Do it', max_marks=50)

        self.scenario_slide = Slide.objects.create(lesson=self.lesson, order=4, slide_type=Slide.SlideType.SCENARIO)
        start_node = ScenarioNode.objects.create(slide=self.scenario_slide, node_key='start', is_start=True)
        end_node = ScenarioNode.objects.create(slide=self.scenario_slide, node_key='end')
        ScenarioChoice.objects.create(node=start_node, choice_text='Go', next_node=end_node, order=1)
        ScenarioChoice.objects.create(node=start_node, choice_text='Stop', next_node=None, order=2)

    def test_platform_admin_can_clone_course_into_organization(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.platform_admin)

        response = self.client.post(f'/api/courses/{self.platform_course.slug}/clone/', {'organization': self.org.id})
        self.assertEqual(response.status_code, 201)

        cloned = Course.objects.get(slug=response.data['slug'])
        self.assertEqual(cloned.content_owner, Course.ContentOwner.ORGANIZATION)
        self.assertEqual(cloned.organization_id, self.org.id)
        self.assertEqual(cloned.cloned_from_id, self.platform_course.id)
        self.assertFalse(cloned.is_published)
        self.assertNotEqual(cloned.id, self.platform_course.id)

        cloned_lesson = Lesson.objects.get(module__course=cloned)
        self.assertEqual(Slide.objects.filter(lesson=cloned_lesson).count(), 4)

        cloned_content_slide = Slide.objects.get(lesson=cloned_lesson, order=1)
        self.assertEqual(cloned_content_slide.elements.count(), 1)
        self.assertEqual(cloned_content_slide.elements.first().rich_text, 'Hello')

        cloned_quiz_slide = Slide.objects.get(lesson=cloned_lesson, order=2)
        cloned_question = Question.objects.get(quiz__slide=cloned_quiz_slide)
        self.assertEqual(cloned_question.buckets.count(), 1)
        self.assertEqual(cloned_question.categorize_items.count(), 1)
        self.assertEqual(cloned_question.categorize_items.first().correct_bucket.question_id, cloned_question.id)
        self.assertEqual(cloned_question.hotspot_regions.count(), 1)
        self.assertEqual(cloned_question.word_bank_tokens.count(), 1)
        self.assertEqual(cloned_question.choices.count(), 1)

        cloned_assignment_slide = Slide.objects.get(lesson=cloned_lesson, order=3)
        self.assertEqual(cloned_assignment_slide.assignment.max_marks, 50)

        cloned_scenario_slide = Slide.objects.get(lesson=cloned_lesson, order=4)
        cloned_start = ScenarioNode.objects.get(slide=cloned_scenario_slide, node_key='start')
        cloned_end = ScenarioNode.objects.get(slide=cloned_scenario_slide, node_key='end')
        self.assertEqual(cloned_start.choices.count(), 2)
        self.assertEqual(cloned_start.choices.get(choice_text='Go').next_node_id, cloned_end.id)
        self.assertIsNone(cloned_start.choices.get(choice_text='Stop').next_node_id)

        # Source course is untouched.
        self.assertEqual(Slide.objects.filter(lesson=self.lesson).count(), 4)
        self.assertFalse(CourseAccess.objects.filter(course=self.platform_course, organization=self.org).exists())

    def test_cannot_clone_an_organization_owned_course(self):
        self.auth_as(self.platform_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/clone/', {'organization': self.other_org.id}
        )
        self.assertEqual(response.status_code, 400)

    def test_org_admin_cannot_clone_platform_course(self):
        self.auth_as(self.org_admin)
        response = self.client.post(f'/api/courses/{self.platform_course.slug}/clone/', {'organization': self.org.id})
        self.assertEqual(response.status_code, 404)


class VideoStreamingTests(BaseAPITestCase):
    """
    Phase 36: an uploaded video Element is served through a short-lived,
    per-user signed streaming URL (courses.video_streaming) rather than its
    raw, permanently-public storage URL.
    """

    def setUp(self):
        super().setUp()
        self.video_slide = Slide.objects.create(
            lesson=self.lesson1, order=98, title='Intro Video', slide_type=Slide.SlideType.CONTENT,
        )
        self.video_content = b'fake-video-bytes-0123456789'
        self.video_element = Element.objects.create(
            slide=self.video_slide,
            order=1,
            element_type=Element.ElementType.VIDEO_AUDIO,
            video_file=SimpleUploadedFile('lesson.mp4', self.video_content, content_type='video/mp4'),
        )

    def _stream_url(self, element, user):
        token = build_video_stream_token(user.id, element.id)
        return f'/api/elements/{element.id}/video/?token={token}'

    def test_element_serializer_returns_streaming_url_not_raw_file_url(self):
        self.auth_as(self.learner)
        response = self.client.get(f'/api/elements/?slide={self.video_slide.id}')
        video_file_url = response.data[0]['video_file']
        self.assertIn(f'/api/elements/{self.video_element.id}/video/', video_file_url)
        self.assertIn('token=', video_file_url)
        self.assertNotIn('/media/element_videos/', video_file_url)

    def test_stream_endpoint_serves_full_video_with_valid_token(self):
        response = self.client.get(self._stream_url(self.video_element, self.learner))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'video/mp4')
        self.assertEqual(b''.join(response.streaming_content), self.video_content)

    def test_stream_endpoint_supports_byte_range_requests(self):
        response = self.client.get(self._stream_url(self.video_element, self.learner), HTTP_RANGE='bytes=5-9')
        self.assertEqual(response.status_code, 206)
        self.assertEqual(response['Content-Range'], f'bytes 5-9/{len(self.video_content)}')
        self.assertEqual(b''.join(response.streaming_content), self.video_content[5:10])

    def test_stream_endpoint_rejects_missing_or_garbage_token(self):
        no_token = self.client.get(f'/api/elements/{self.video_element.id}/video/')
        self.assertEqual(no_token.status_code, 403)

        bad_token = self.client.get(f'/api/elements/{self.video_element.id}/video/?token=garbage')
        self.assertEqual(bad_token.status_code, 403)

    def test_stream_endpoint_rejects_token_minted_for_a_different_element(self):
        other_element = Element.objects.create(
            slide=self.video_slide, order=2, element_type=Element.ElementType.TEXT, rich_text='<p>x</p>',
        )
        token = build_video_stream_token(self.learner.id, other_element.id)
        response = self.client.get(f'/api/elements/{self.video_element.id}/video/?token={token}')
        self.assertEqual(response.status_code, 403)

    def test_stream_endpoint_rejects_user_outside_course_organization(self):
        response = self.client.get(self._stream_url(self.video_element, self.other_org_learner))
        self.assertEqual(response.status_code, 404)

    def test_stream_endpoint_respects_demo_lesson_lock(self):
        demo_learner = User.objects.create_user(
            email='demo-video@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org, is_demo=True,
        )
        self.published_org_course.is_demo_available = True
        self.published_org_course.save()
        # lesson1 (the video's lesson) is intentionally left without a
        # DemoLessonAccess grant, so it stays locked for this demo user.

        response = self.client.get(self._stream_url(self.video_element, demo_learner))
        self.assertEqual(response.status_code, 403)

        DemoLessonAccess.objects.create(course=self.published_org_course, lesson=self.lesson1)
        granted_response = self.client.get(self._stream_url(self.video_element, demo_learner))
        self.assertEqual(granted_response.status_code, 200)


class ModuleLessonBuilderTests(BaseAPITestCase):
    def test_org_admin_can_create_module_and_lesson(self):
        self.auth_as(self.org_admin)
        module_response = self.client.post('/api/modules/', {
            'course': self.published_org_course.id, 'title': 'New Module', 'order': 5,
        })
        self.assertEqual(module_response.status_code, 201)

        lesson_response = self.client.post('/api/lessons/', {
            'module': module_response.data['id'], 'title': 'New Lesson',
            'lesson_type': 'TEXT', 'order': 1,
        })
        self.assertEqual(lesson_response.status_code, 201)

    def test_instructor_cannot_create_module_for_other_org_course(self):
        self.auth_as(self.instructor)
        response = self.client.post('/api/modules/', {
            'course': self.other_org_course.id, 'title': 'Nope', 'order': 1,
        })
        self.assertEqual(response.status_code, 400)

    def test_lesson_file_extension_validated_through_api(self):
        self.auth_as(self.org_admin)
        bad_file = SimpleUploadedFile('lesson.exe', b'not a video', content_type='application/octet-stream')
        response = self.client.post('/api/lessons/', {
            'module': self.module.id, 'title': 'Bad Video', 'lesson_type': 'VIDEO',
            'order': 9, 'content_file': bad_file,
        }, format='multipart')
        self.assertEqual(response.status_code, 400)

    def test_learner_cannot_create_lesson(self):
        self.auth_as(self.learner)
        response = self.client.post('/api/lessons/', {
            'module': self.module.id, 'title': 'Nope', 'lesson_type': 'TEXT', 'order': 9,
        })
        self.assertEqual(response.status_code, 403)


class QuizBuilderTests(BaseAPITestCase):
    def test_org_admin_can_build_quiz_question_choice(self):
        self.auth_as(self.org_admin)
        new_quiz_slide = Slide.objects.create(
            lesson=self.lesson2, order=99, title='New Quiz', slide_type=Slide.SlideType.QUIZ,
        )
        quiz_response = self.client.post('/api/quizzes/', {
            'slide': new_quiz_slide.id, 'title': 'New Quiz', 'pass_percentage': 60,
        })
        self.assertEqual(quiz_response.status_code, 201)

        question_response = self.client.post('/api/questions/', {
            'quiz': quiz_response.data['id'], 'question_text': 'Q1?',
            'question_type': 'SINGLE_CHOICE', 'order': 1, 'points': 1,
        })
        self.assertEqual(question_response.status_code, 201)

        choice_response = self.client.post('/api/choices/', {
            'question': question_response.data['id'], 'choice_text': 'A', 'is_correct': True,
        })
        self.assertEqual(choice_response.status_code, 201)

    def test_learner_cannot_create_quiz(self):
        self.auth_as(self.learner)
        new_quiz_slide = Slide.objects.create(
            lesson=self.lesson2, order=99, title='Nope', slide_type=Slide.SlideType.QUIZ,
        )
        response = self.client.post('/api/quizzes/', {
            'slide': new_quiz_slide.id, 'title': 'Nope',
        })
        self.assertEqual(response.status_code, 403)

    def test_cannot_add_question_to_other_org_quiz(self):
        other_module = Module.objects.create(course=self.other_org_course, title='Other', order=1)
        other_lesson = Lesson.objects.create(module=other_module, title='Other lesson', order=1)
        other_quiz_slide = Slide.objects.create(
            lesson=other_lesson, order=1, title='Other Quiz', slide_type=Slide.SlideType.QUIZ,
        )
        other_quiz = Quiz.objects.create(slide=other_quiz_slide, title='Other Quiz')
        self.auth_as(self.instructor)
        response = self.client.post('/api/questions/', {
            'quiz': other_quiz.id, 'question_text': 'Q?', 'order': 1, 'points': 1,
        })
        self.assertEqual(response.status_code, 400)


class BulkEnrollTests(BaseAPITestCase):
    def test_bulk_enroll_csv_enrolls_existing_users(self):
        csv_content = f'email\n{self.learner.email}\nnobody@example.com\n'.encode()
        upload = SimpleUploadedFile('emails.csv', csv_content, content_type='text/csv')

        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['enrolled'], [self.learner.email])
        self.assertEqual(response.data['not_found'], ['nobody@example.com'])
        self.assertTrue(Enrollment.objects.filter(user=self.learner, course=self.published_org_course).exists())

    def test_bulk_enroll_already_enrolled_is_reported_separately(self):
        Enrollment.objects.create(user=self.learner, course=self.published_org_course)
        csv_content = f'{self.learner.email}\n'.encode()
        upload = SimpleUploadedFile('emails.csv', csv_content, content_type='text/csv')

        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.data['already_enrolled'], [self.learner.email])
        self.assertEqual(response.data['enrolled'], [])

    def test_bulk_enroll_forbidden_for_learner(self):
        upload = SimpleUploadedFile('emails.csv', b'a@example.com\n', content_type='text/csv')
        self.auth_as(self.learner)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.status_code, 403)

    def test_bulk_enroll_rejects_course_outside_org(self):
        upload = SimpleUploadedFile('emails.csv', f'{self.learner.email}\n'.encode(), content_type='text/csv')
        self.auth_as(self.instructor)
        response = self.client.post(
            f'/api/courses/{self.other_org_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.status_code, 404)

    def test_bulk_enroll_rejects_user_from_a_different_organization(self):
        csv_content = f'{self.other_org_learner.email}\n'.encode()
        upload = SimpleUploadedFile('emails.csv', csv_content, content_type='text/csv')

        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['wrong_organization'], [self.other_org_learner.email])
        self.assertEqual(response.data['enrolled'], [])
        self.assertFalse(
            Enrollment.objects.filter(user=self.other_org_learner, course=self.published_org_course).exists()
        )

    def test_bulk_enroll_platform_admin_may_enroll_any_org_into_a_platform_course(self):
        csv_content = f'{self.other_org_learner.email}\n'.encode()
        upload = SimpleUploadedFile('emails.csv', csv_content, content_type='text/csv')

        self.auth_as(self.platform_admin)
        response = self.client.post(
            f'/api/courses/{self.platform_course.slug}/bulk-enroll/', {'file': upload}, format='multipart'
        )
        self.assertEqual(response.data['enrolled'], [self.other_org_learner.email])
        self.assertTrue(Enrollment.objects.filter(user=self.other_org_learner, course=self.platform_course).exists())


class InviteLearnerTests(BaseAPITestCase):
    def test_invite_enrolls_existing_user_in_own_org(self):
        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/invite/', {'email': self.learner.email}
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(Enrollment.objects.filter(user=self.learner, course=self.published_org_course).exists())

    def test_invite_rejects_user_from_a_different_organization(self):
        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/invite/', {'email': self.other_org_learner.email}
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            Enrollment.objects.filter(user=self.other_org_learner, course=self.published_org_course).exists()
        )

    def test_invite_unknown_email_returns_404(self):
        self.auth_as(self.org_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/invite/', {'email': 'nobody@example.com'}
        )
        self.assertEqual(response.status_code, 404)

    def test_invite_rejects_course_outside_org(self):
        self.auth_as(self.instructor)
        response = self.client.post(
            f'/api/courses/{self.other_org_course.slug}/invite/', {'email': self.learner.email}
        )
        self.assertEqual(response.status_code, 404)


class EnrollmentReportTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.enrollment = Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(
            user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=90,
        )
        Enrollment.objects.create(user=self.other_org_learner, course=self.other_org_course)

    def test_learner_forbidden_from_report(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/reports/enrollments/')
        self.assertEqual(response.status_code, 403)

    def test_org_admin_sees_only_their_org_rows(self):
        self.auth_as(self.org_admin)
        response = self.client.get('/api/reports/enrollments/')
        self.assertEqual(len(response.data), 1)
        row = response.data[0]
        self.assertEqual(row['learner_email'], self.learner.email)
        self.assertEqual(row['status'], 'COMPLETED')
        self.assertEqual(row['score_percent'], 90.0)

    def test_platform_admin_sees_all_rows(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/enrollments/')
        self.assertEqual(len(response.data), 2)

    def test_report_filterable_by_status(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/enrollments/?status=NOT_STARTED')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['learner_email'], self.other_org_learner.email)

    def test_report_csv_export(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/enrollments/?export=csv')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        content = response.content.decode()
        self.assertIn(self.learner.email, content)
        self.assertIn('90.0', content)

    def test_report_csv_export_neutralizes_formula_injection(self):
        evil_learner = User.objects.create_user(
            email='formula@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org,
            first_name='=HYPERLINK("http://evil.test")', last_name='X',
        )
        Enrollment.objects.create(user=evil_learner, course=self.published_org_course)

        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/enrollments/?export=csv')
        content = response.content.decode()
        self.assertNotIn('\n=HYPERLINK', content)
        self.assertIn("'=HYPERLINK", content)


class AdminAnalyticsTests(BaseAPITestCase):
    """Phase 37: admin analytics dashboard grouped by Organization."""

    def setUp(self):
        super().setUp()
        # Second quiz on lesson2 so final_score can diverge from any single
        # quiz's own pass/fail, same fixture shape as
        # CourseAverageCertificateEligibilityTests.
        self.quiz2_slide = Slide.objects.create(
            lesson=self.lesson2, order=99, title='Second Exam', slide_type=Slide.SlideType.QUIZ,
        )
        self.quiz2 = Quiz.objects.create(slide=self.quiz2_slide, title='Second Exam', pass_percentage=70)

        self.enrollment = Enrollment.objects.create(
            user=self.learner, course=self.published_org_course,
            status=Enrollment.Status.COMPLETED, progress_percent=100,
        )
        SlideProgress.objects.create(
            enrollment=self.enrollment, slide=self.quiz_slide, time_spent_seconds=300, completed_at=timezone.now(),
        )
        SlideProgress.objects.create(
            enrollment=self.enrollment, slide=self.quiz2_slide, time_spent_seconds=120, completed_at=timezone.now(),
        )
        # quiz1: failed then retaken and passed; quiz2: passed on the first try.
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=False, score_percent=40)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=2, passed=True, score_percent=90)
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz2, attempt_number=1, passed=True, score_percent=100)
        # Average of best scores = (90 + 100) / 2 = 95, clears the 70% default threshold.

        self.other_org_enrollment = Enrollment.objects.create(
            user=self.other_org_learner, course=self.other_org_course, status=Enrollment.Status.IN_PROGRESS,
            progress_percent=40,
        )

    def _row_for(self, response, user_email):
        for org_group in response.data:
            for row in org_group['rows']:
                if row['user_email'] == user_email:
                    return row
        return None

    def test_learner_forbidden(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/reports/analytics/')
        self.assertEqual(response.status_code, 403)

    def test_org_admin_sees_only_their_own_organization_grouped(self):
        self.auth_as(self.org_admin)
        response = self.client.get('/api/reports/analytics/')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['organization_name'], 'Acme Bank')
        self.assertEqual(len(response.data[0]['rows']), 1)

    def test_org_admin_organization_filter_param_is_ignored(self):
        self.auth_as(self.org_admin)
        response = self.client.get(f'/api/reports/analytics/?organization={self.other_org.id}')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['organization_name'], 'Acme Bank')

    def test_platform_admin_sees_every_organization_grouped(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/')
        org_names = {group['organization_name'] for group in response.data}
        self.assertEqual(org_names, {'Acme Bank', 'Other Bank'})

    def test_platform_admin_can_filter_by_organization(self):
        self.auth_as(self.platform_admin)
        response = self.client.get(f'/api/reports/analytics/?organization={self.other_org.id}')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['organization_name'], 'Other Bank')

    def test_course_filter_narrows_rows(self):
        self.auth_as(self.platform_admin)
        response = self.client.get(f'/api/reports/analytics/?course={self.other_org_course.id}')
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['rows'][0]['user_email'], self.other_org_learner.email)

    def test_row_reports_completion_pass_status_time_spent_and_final_score(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/')
        row = self._row_for(response, self.learner.email)
        self.assertEqual(row['progress_percent'], 100)
        self.assertEqual(row['pass_status'], 'PASSED')
        self.assertEqual(row['final_score'], 95.0)
        self.assertEqual(row['time_spent_seconds'], 420)
        self.assertEqual(row['total_quiz_attempts'], 3)

    def test_row_includes_per_quiz_retake_attempts_and_scores(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/')
        row = self._row_for(response, self.learner.email)
        quiz1 = next(q for q in row['quizzes'] if q['quiz_id'] == self.quiz.id)
        self.assertEqual(quiz1['attempt_count'], 2)
        self.assertEqual([a['score_percent'] for a in quiz1['attempts']], [40.0, 90.0])
        self.assertEqual(quiz1['best_score'], 90.0)

    def test_incomplete_enrollment_reports_status_not_a_pass_fail_verdict(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/')
        row = self._row_for(response, self.other_org_learner.email)
        self.assertEqual(row['pass_status'], 'IN_PROGRESS')
        self.assertIsNone(row['final_score'])

    def test_average_below_threshold_reports_failed_despite_individual_pass(self):
        self.other_org_learner.organization = self.org
        self.other_org_learner.save()
        failing_module = Module.objects.create(course=self.unpublished_org_course, title='Only Module', order=1)
        failing_lesson = Lesson.objects.create(module=failing_module, title='Only Lesson', order=1, estimated_minutes=5)
        failing_quiz_slide = Slide.objects.create(
            lesson=failing_lesson, order=1, title='Only Exam', slide_type=Slide.SlideType.QUIZ,
        )
        failing_quiz = Quiz.objects.create(slide=failing_quiz_slide, title='Only Exam', pass_percentage=50)
        QuizAttempt.objects.create(
            user=self.other_org_learner, quiz=failing_quiz, attempt_number=1, passed=True, score_percent=60,
        )
        Enrollment.objects.create(
            user=self.other_org_learner, course=self.unpublished_org_course,
            status=Enrollment.Status.COMPLETED, progress_percent=100,
        )

        self.auth_as(self.platform_admin)
        response = self.client.get(f'/api/reports/analytics/?course={self.unpublished_org_course.id}')
        row = self._row_for(response, self.other_org_learner.email)
        self.assertEqual(row['pass_status'], 'FAILED')
        self.assertEqual(row['final_score'], 60.0)

    def test_xlsx_export_matches_on_screen_data(self):
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/?export=xlsx')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        header_row = [cell.value for cell in sheet[1]]
        self.assertEqual(
            header_row,
            [
                'Organization', 'User Name', 'User Email', 'Course', 'Status',
                '% Completion', 'Pass/Fail', 'Final Score', 'Time Spent',
                'Total Quiz Attempts', 'Attempt Details',
            ],
        )

        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        learner_row = next(r for r in data_rows if r[2] == self.learner.email)
        self.assertEqual(learner_row[0], 'Acme Bank')
        self.assertEqual(learner_row[4], 'COMPLETED')
        self.assertEqual(learner_row[5], 1.0)  # 100% stored as a fraction for the '0%' number format
        self.assertEqual(learner_row[6], 'PASSED')
        self.assertEqual(learner_row[7], 95.0)
        self.assertEqual(learner_row[8], '7m')
        self.assertEqual(learner_row[9], 3)
        self.assertIn('Final Exam', learner_row[10])
        self.assertIn('Second Exam', learner_row[10])

    def test_xlsx_export_neutralizes_formula_injection(self):
        evil_learner = User.objects.create_user(
            email='evil-analytics@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org,
            first_name='=HYPERLINK("http://evil.test")', last_name='X',
        )
        Enrollment.objects.create(user=evil_learner, course=self.published_org_course)

        self.auth_as(self.platform_admin)
        response = self.client.get('/api/reports/analytics/?export=xlsx')
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        names = [row[1] for row in sheet.iter_rows(min_row=2, values_only=True)]
        self.assertIn("'=HYPERLINK(\"http://evil.test\") X", names)


class RateLimitingTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()

    def tearDown(self):
        cache.clear()
        super().tearDown()

    def test_login_is_rate_limited(self):
        payload = {'email': self.learner.email, 'password': 'wrong-password'}
        statuses = [self.client.post('/api/auth/login/', payload).status_code for _ in range(10)]
        self.assertTrue(all(code == 401 for code in statuses))

        throttled = self.client.post('/api/auth/login/', payload)
        self.assertEqual(throttled.status_code, 429)

    def test_quiz_submit_is_rate_limited(self):
        self.auth_as(self.learner)
        payload = {'answers': [
            {'question': self.q1.id, 'selected_choices': [self.q1_right.id]},
            {'question': self.q2.id, 'selected_choices': [self.q2_right.id]},
        ]}
        # max_attempts=2 on self.quiz would mask the throttle after 2 tries, so
        # raise it for this test to isolate the rate limit specifically.
        self.quiz.max_attempts = None
        self.quiz.save()

        statuses = [
            self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json').status_code
            for _ in range(20)
        ]
        self.assertTrue(all(code == 201 for code in statuses))

        throttled = self.client.post(f'/api/quizzes/{self.quiz.id}/submit/', payload, format='json')
        self.assertEqual(throttled.status_code, 429)


class AuditLogTests(BaseAPITestCase):
    def test_login_is_audit_logged(self):
        self.client.post('/api/auth/login/', {'email': self.learner.email, 'password': 'pass12345'})
        log = AuditLog.objects.get(action=AuditLog.Action.LOGIN, object_id=str(self.learner.id))
        self.assertEqual(log.user, self.learner)

    def test_failed_login_is_not_audit_logged(self):
        self.client.post('/api/auth/login/', {'email': self.learner.email, 'password': 'wrong'})
        self.assertFalse(AuditLog.objects.filter(action=AuditLog.Action.LOGIN).exists())

    def test_course_creation_is_audit_logged(self):
        self.auth_as(self.platform_admin)
        response = self.client.post('/api/courses/', {'title': 'Audited Course', 'slug': 'audited-course'})
        log = AuditLog.objects.get(action=AuditLog.Action.COURSE_CREATED, object_id=str(response.data['id']))
        self.assertEqual(log.user, self.platform_admin)

    def test_enrollment_create_and_update_are_audit_logged(self):
        self.auth_as(self.learner)
        create_response = self.client.post('/api/enrollments/', {'course': self.published_org_course.id})
        enrollment_id = create_response.data['id']
        self.assertTrue(
            AuditLog.objects.filter(action=AuditLog.Action.ENROLLMENT_CREATED, object_id=str(enrollment_id)).exists()
        )

        self.client.patch(f'/api/enrollments/{enrollment_id}/', {'status': 'IN_PROGRESS'})
        self.assertTrue(
            AuditLog.objects.filter(action=AuditLog.Action.ENROLLMENT_UPDATED, object_id=str(enrollment_id)).exists()
        )

    def test_certificate_generation_is_audit_logged(self):
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course, status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)
        certificate = generate_certificate(self.learner, self.published_org_course)

        log = AuditLog.objects.get(action=AuditLog.Action.CERTIFICATE_GENERATED, object_id=str(certificate.id))
        self.assertEqual(log.user, self.learner)


class CertificateTemplateRenderingTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course, status=Enrollment.Status.COMPLETED,
        )
        QuizAttempt.objects.create(user=self.learner, quiz=self.quiz, attempt_number=1, passed=True, score_percent=100)

    def test_generate_certificate_renders_a_valid_pdf_against_seeded_default_template(self):
        certificate = generate_certificate(self.learner, self.published_org_course)
        pdf_bytes = certificate.pdf_file.read()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))

    def test_generate_certificate_raises_when_no_template_is_configured(self):
        CertificateTemplate.objects.all().delete()
        with self.assertRaises(CertificateIssuanceError):
            generate_certificate(self.learner, self.published_org_course)

    def test_course_specific_template_overrides_platform_default(self):
        course_template = make_test_certificate_template(name='Course-specific template')
        self.published_org_course.certificate_template = course_template
        self.published_org_course.save()

        certificate = generate_certificate(self.learner, self.published_org_course)
        pdf_bytes = certificate.pdf_file.read()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))

    def test_learners_organization_template_used_when_no_course_override_exists(self):
        make_test_certificate_template(name='Acme template', organization=self.org)

        certificate = generate_certificate(self.learner, self.published_org_course)

        self.assertEqual(certificate.pdf_file.read()[:4], b'%PDF')
        # Confirms the org template — not the seeded platform default — was
        # actually the one resolved, not just that *a* PDF rendered.
        self.assertEqual(_resolve_template(self.published_org_course, self.learner).organization_id, self.org.id)

    def test_course_override_still_wins_over_the_learners_organization_template(self):
        make_test_certificate_template(name='Acme template', organization=self.org)
        course_template = make_test_certificate_template(name='Course-specific template')
        self.published_org_course.certificate_template = course_template
        self.published_org_course.save()

        resolved = _resolve_template(self.published_org_course, self.learner)
        self.assertEqual(resolved.id, course_template.id)

    def test_platform_default_used_when_learner_has_no_organization(self):
        self.learner.organization = None
        self.learner.save()

        resolved = _resolve_template(self.published_org_course, self.learner)
        self.assertTrue(resolved.is_default)

    def test_saving_a_new_default_template_unsets_the_previous_one(self):
        original_default = CertificateTemplate.objects.get(is_default=True)
        new_default = make_test_certificate_template(name='New default', is_default=True)

        original_default.refresh_from_db()
        self.assertFalse(original_default.is_default)
        self.assertTrue(new_default.is_default)
        self.assertEqual(CertificateTemplate.objects.filter(is_default=True).count(), 1)

    def test_long_name_and_course_title_render_without_error(self):
        self.learner.first_name = 'Alexandria' * 5
        self.learner.last_name = 'Featherington-Papadopoulos' * 3
        self.learner.save()
        self.published_org_course.title = 'Advanced Regulatory Compliance and Risk Management ' * 4
        self.published_org_course.save()

        certificate = generate_certificate(self.learner, self.published_org_course)
        pdf_bytes = certificate.pdf_file.read()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))


class FontAutoShrinkTests(TestCase):
    """Unit-level coverage of the name/course-title overflow guard in certificates.services._fit_font."""

    def setUp(self):
        self.image = Image.new('RGB', (2000, 1414), color='white')
        self.draw = ImageDraw.Draw(self.image)
        self.max_width_px = 2000 * 0.84

    def test_short_text_keeps_the_requested_font_size(self):
        font = _fit_font(None, 60, 'Jane Doe', self.draw, self.max_width_px)
        self.assertEqual(font.size, 60)

    def test_long_text_shrinks_down_to_the_minimum_font_size(self):
        # A max width this narrow can never be satisfied, so the loop should bottom out at the floor.
        font = _fit_font(None, 110, 'A very long staff name that will not fit', self.draw, max_width_px=50)
        self.assertEqual(font.size, MIN_AUTO_SHRINK_FONT_SIZE)

    def test_moderately_long_text_shrinks_below_the_initial_size(self):
        font = _fit_font(None, 110, 'A very long staff name that will not fit ' * 3, self.draw, self.max_width_px)
        self.assertLess(font.size, 110)

    def test_shrunk_text_fits_within_the_max_width(self):
        text = 'A Notably Long Course Title That Should Trigger Auto-Shrink Handling'
        font = _fit_font(None, 110, text, self.draw, self.max_width_px)
        left, _top, right, _bottom = self.draw.textbbox((0, 0), text, font=font)
        self.assertTrue((right - left) <= self.max_width_px or font.size == MIN_AUTO_SHRINK_FONT_SIZE)


class CertificateTemplateViewSetTests(BaseAPITestCase):
    def test_learner_cannot_list_certificate_templates(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/certificate-templates/')
        self.assertEqual(response.status_code, 403)

    def test_instructor_cannot_see_the_platform_default_template(self):
        # Org-scoped now: the platform-level template is a platform-wide
        # branding asset, not an individual organization's to see or manage.
        self.auth_as(self.instructor)
        response = self.client.get('/api/certificate-templates/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 0)

    def test_instructor_sees_and_can_update_only_their_own_organizations_template(self):
        own_template = make_test_certificate_template(name='Acme template', organization=self.org)
        make_test_certificate_template(name='Other Bank template', organization=self.other_org)

        self.auth_as(self.instructor)
        listing = self.client.get('/api/certificate-templates/')
        self.assertEqual(listing.status_code, 200)
        self.assertEqual([t['id'] for t in listing.data], [own_template.id])

        response = self.client.patch(
            f'/api/certificate-templates/{own_template.id}/',
            {'staff_name_x_percent': 42.5, 'staff_name_text_align': 'LEFT'},
            format='json',
        )
        self.assertEqual(response.status_code, 200)
        own_template.refresh_from_db()
        self.assertEqual(own_template.staff_name_x_percent, 42.5)

    def test_instructor_cannot_access_another_organizations_template(self):
        other_template = make_test_certificate_template(name='Other Bank template', organization=self.other_org)
        self.auth_as(self.instructor)
        response = self.client.get(f'/api/certificate-templates/{other_template.id}/')
        self.assertEqual(response.status_code, 404)

    def test_platform_admin_sees_every_organizations_template_and_the_platform_default(self):
        make_test_certificate_template(name='Acme template', organization=self.org)
        self.auth_as(self.platform_admin)
        response = self.client.get('/api/certificate-templates/')
        self.assertEqual(response.status_code, 200)
        # The seeded platform default (organization=None) plus the one just created.
        self.assertEqual(len(response.data), 2)

    def test_instructor_can_create_a_template_for_their_own_organization(self):
        image_buffer = io.BytesIO()
        Image.new('RGB', (400, 300), color='white').save(image_buffer, format='PNG')
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/certificate-templates/',
            {
                'name': 'Acme certificate',
                'organization': self.org.id,
                'background_image': SimpleUploadedFile('bg.png', image_buffer.getvalue(), content_type='image/png'),
            },
            format='multipart',
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['organization'], self.org.id)
        self.assertFalse(response.data['is_default'])

    def test_instructor_cannot_create_a_template_for_a_different_organization(self):
        image_buffer = io.BytesIO()
        Image.new('RGB', (400, 300), color='white').save(image_buffer, format='PNG')
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/certificate-templates/',
            {
                'name': 'Sneaky',
                'organization': self.other_org.id,
                'background_image': SimpleUploadedFile('bg.png', image_buffer.getvalue(), content_type='image/png'),
            },
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)

    def test_instructor_cannot_create_the_platform_default_template(self):
        image_buffer = io.BytesIO()
        Image.new('RGB', (400, 300), color='white').save(image_buffer, format='PNG')
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/certificate-templates/',
            {'name': 'Sneaky platform template', 'background_image': SimpleUploadedFile('bg.png', image_buffer.getvalue(), content_type='image/png')},
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)

    def test_cannot_mark_an_organization_scoped_template_as_default(self):
        image_buffer = io.BytesIO()
        Image.new('RGB', (400, 300), color='white').save(image_buffer, format='PNG')
        self.auth_as(self.platform_admin)
        response = self.client.post(
            '/api/certificate-templates/',
            {
                'name': 'Bad combo',
                'organization': self.org.id,
                'is_default': True,
                'background_image': SimpleUploadedFile('bg.png', image_buffer.getvalue(), content_type='image/png'),
            },
            format='multipart',
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('is_default', response.data)

    def test_only_one_template_allowed_per_organization(self):
        make_test_certificate_template(name='First', organization=self.org)
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                make_test_certificate_template(name='Second', organization=self.org)

    def test_unauthenticated_request_is_rejected(self):
        self.client.credentials()
        response = self.client.get('/api/certificate-templates/')
        self.assertEqual(response.status_code, 401)


class DemoUserProvisioningServiceTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(name='Acme Bank', slug='acme-bank-demo')

    def test_creates_demo_learner_with_must_reset_password_and_sends_invite(self):
        user = provision_demo_user(name='Dana Demo', email='dana@example.com', organization=self.org)

        self.assertEqual(user.role, User.Role.LEARNER)
        self.assertTrue(user.is_demo)
        self.assertTrue(user.must_reset_password)
        self.assertEqual(user.first_name, 'Dana')
        self.assertEqual(user.last_name, 'Demo')
        self.assertEqual(user.organization, self.org)

        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(user.email, mail.outbox[0].to)
        self.assertIn('temporary password', mail.outbox[0].body.lower())

    def test_duplicate_email_is_rejected_case_insensitively(self):
        provision_demo_user(name='Dana Demo', email='dana@example.com', organization=self.org)
        with self.assertRaises(UserProvisioningError):
            provision_demo_user(name='Dana Two', email='DANA@example.com', organization=self.org)

    def test_missing_name_is_rejected(self):
        with self.assertRaises(UserProvisioningError):
            provision_demo_user(name='  ', email='dana@example.com', organization=self.org)

    def test_email_send_failure_rolls_back_the_account(self):
        with patch('accounts.services.EmailMultiAlternatives.send', side_effect=Exception('smtp down')):
            with self.assertRaises(UserProvisioningError):
                provision_demo_user(name='Dana Demo', email='dana@example.com', organization=self.org)

        # The whole operation is atomic — a failed invite must not leave a
        # stranded account with a password nobody received.
        self.assertFalse(User.objects.filter(email='dana@example.com').exists())


class DemoUserApiTests(BaseAPITestCase):
    def test_learner_cannot_create_demo_user(self):
        self.auth_as(self.learner)
        response = self.client.post(
            '/api/demo-users/', {'name': 'Dana Demo', 'email': 'dana@example.com', 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(len(mail.outbox), 0)

    def test_admin_can_create_demo_user(self):
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/demo-users/', {'name': 'Dana Demo', 'email': 'dana@example.com', 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(
            User.objects.filter(email='dana@example.com', is_demo=True, must_reset_password=True).exists()
        )
        self.assertEqual(len(mail.outbox), 1)

    def test_duplicate_email_returns_400_not_500(self):
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/demo-users/', {'name': 'Dana Demo', 'email': self.learner.email, 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 400)

    def test_bulk_upload_reports_per_row_success_and_failure(self):
        csv_content = (
            'name,email,organization\n'
            f'Alice Alpha,alice@example.com,{self.org.name}\n'
            f'Bob Beta,bob@example.com,{self.org.name}\n'
            f'Bob Duplicate,bob@example.com,{self.org.name}\n'
            'Missing Org,noorg@example.com,Nonexistent Org\n'
            'OnlyTwoColumns,twocols@example.com\n'
        )
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], ['alice@example.com', 'bob@example.com'])

        failures_by_email = {f['email']: f['reason'] for f in response.data['failed']}
        self.assertIn('Duplicate email', failures_by_email['bob@example.com'])
        self.assertIn('not found', failures_by_email['noorg@example.com'])
        self.assertIn('Malformed row', failures_by_email['twocols@example.com'])
        self.assertEqual(len(mail.outbox), 2)

    def test_bulk_upload_captures_designation_and_phone_number(self):
        csv_content = (
            'name,email,organization,designation,phone_number\n'
            f'Alice Alpha,alice@example.com,{self.org.name},Compliance Officer,+977-1-4123456\n'
        )
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], ['alice@example.com'])

        user = User.objects.get(email='alice@example.com')
        self.assertEqual(user.designation, 'Compliance Officer')
        self.assertEqual(user.phone_number, '+977-1-4123456')

    def test_bulk_upload_extended_columns_captures_titles_and_assessment_level(self):
        csv_content = (
            'Name,Email,Corporate Title,Functional Title,Branch/Department,Assessment Level,Organization\n'
            f'Alice Alpha,alice@example.com,VP,Compliance Analyst,Head Office,Officer,{self.org.name}\n'
            f'Bob Beta,bob@example.com,SVP,Risk Lead,Branch Office,Senior Management,{self.org.name}\n'
        )
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], ['alice@example.com', 'bob@example.com'])

        alice = User.objects.get(email='alice@example.com')
        self.assertEqual(alice.corporate_title, 'VP')
        self.assertEqual(alice.functional_title, 'Compliance Analyst')
        self.assertEqual(alice.branch_department, 'Head Office')
        self.assertEqual(alice.assessment_level, 'officer')

        bob = User.objects.get(email='bob@example.com')
        self.assertEqual(bob.assessment_level, 'senior_management')

    def test_bulk_upload_extended_columns_rejects_invalid_assessment_level(self):
        csv_content = (
            'Name,Email,Corporate Title,Functional Title,Branch/Department,Assessment Level,Organization\n'
            f'Alice Alpha,alice@example.com,VP,Compliance Analyst,Head Office,Director,{self.org.name}\n'
        )
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertEqual(len(response.data['failed']), 1)
        self.assertIn('Invalid Assessment Level', response.data['failed'][0]['reason'])
        self.assertFalse(User.objects.filter(email='alice@example.com').exists())

    def test_bulk_upload_extended_columns_rejects_missing_assessment_level(self):
        csv_content = (
            'Name,Email,Corporate Title,Functional Title,Branch/Department,Assessment Level,Organization\n'
            f'Alice Alpha,alice@example.com,VP,Compliance Analyst,Head Office,,{self.org.name}\n'
        )
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertEqual(len(response.data['failed']), 1)
        self.assertIn('Invalid Assessment Level', response.data['failed'][0]['reason'])

    def test_bulk_upload_reports_existing_user_as_failure_not_500(self):
        csv_content = f'name,email,organization\nExisting User,{self.learner.email},{self.org.name}\n'
        upload = SimpleUploadedFile('demo_users.csv', csv_content.encode('utf-8'), content_type='text/csv')

        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertEqual(len(response.data['failed']), 1)
        self.assertIn('already exists', response.data['failed'][0]['reason'])

    def test_bulk_upload_requires_a_file(self):
        self.auth_as(self.instructor)
        response = self.client.post('/api/demo-users/bulk/', {}, format='multipart')
        self.assertEqual(response.status_code, 400)

    def test_learner_cannot_bulk_upload(self):
        upload = SimpleUploadedFile('demo_users.csv', b'name,email,organization\n', content_type='text/csv')
        self.auth_as(self.learner)
        response = self.client.post('/api/demo-users/bulk/', {'file': upload}, format='multipart')
        self.assertEqual(response.status_code, 403)


class OrgAdminApiTests(BaseAPITestCase):
    def test_platform_admin_can_create_org_admin(self):
        self.auth_as(self.platform_admin)
        response = self.client.post(
            '/api/org-admins/', {'name': 'Jane Manager', 'email': 'jane@example.com', 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['role'], 'ORG_ADMIN')
        self.assertFalse(response.data['is_demo'])
        self.assertTrue(response.data['must_reset_password'])

        user = User.objects.get(email='jane@example.com')
        self.assertEqual(user.role, User.Role.ORG_ADMIN)
        self.assertFalse(user.is_demo)
        self.assertEqual(user.organization, self.org)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn('administrator', mail.outbox[0].subject.lower())

    def test_org_admin_cannot_create_org_admin(self):
        self.auth_as(self.org_admin)
        response = self.client.post(
            '/api/org-admins/', {'name': 'Jane Manager', 'email': 'jane@example.com', 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(len(mail.outbox), 0)

    def test_instructor_cannot_create_org_admin(self):
        self.auth_as(self.instructor)
        response = self.client.post(
            '/api/org-admins/', {'name': 'Jane Manager', 'email': 'jane@example.com', 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 403)

    def test_duplicate_email_returns_400_not_500(self):
        self.auth_as(self.platform_admin)
        response = self.client.post(
            '/api/org-admins/', {'name': 'Dup', 'email': self.learner.email, 'organization': self.org.id}
        )
        self.assertEqual(response.status_code, 400)


class SetPasswordApiTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.temp_password = 'Temp-Pass-9x7Q'
        self.demo_user = User.objects.create_user(
            email='dana@example.com', password=self.temp_password,
            role=User.Role.LEARNER, organization=self.org,
            first_name='Dana', last_name='Demo',
            is_demo=True, must_reset_password=True,
        )

    def test_new_password_clears_must_reset_flag(self):
        # No current_password field — reaching this endpoint already required
        # a valid access token, which the caller could only have obtained by
        # authenticating with the temp password moments earlier.
        self.auth_as(self.demo_user)
        response = self.client.post('/api/auth/set-password/', {
            'new_password': 'BrandNewPassw0rd1',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data['must_reset_password'])

        self.demo_user.refresh_from_db()
        self.assertFalse(self.demo_user.must_reset_password)
        self.assertTrue(self.demo_user.check_password('BrandNewPassw0rd1'))

    def test_too_short_new_password_is_rejected_by_validators(self):
        self.auth_as(self.demo_user)
        response = self.client.post('/api/auth/set-password/', {'new_password': 'ab1'})
        self.assertEqual(response.status_code, 400)

        self.demo_user.refresh_from_db()
        self.assertTrue(self.demo_user.must_reset_password)

    def test_new_password_missing_a_digit_is_rejected(self):
        self.auth_as(self.demo_user)
        response = self.client.post('/api/auth/set-password/', {'new_password': 'allletters'})
        self.assertEqual(response.status_code, 400)

        self.demo_user.refresh_from_db()
        self.assertTrue(self.demo_user.must_reset_password)

    def test_new_password_missing_a_letter_is_rejected(self):
        self.auth_as(self.demo_user)
        # Also all-numeric, so this doubles as coverage for NumericPasswordValidator.
        response = self.client.post('/api/auth/set-password/', {'new_password': '12345678'})
        self.assertEqual(response.status_code, 400)

        self.demo_user.refresh_from_db()
        self.assertTrue(self.demo_user.must_reset_password)

    def test_me_endpoint_reflects_must_reset_password(self):
        self.auth_as(self.demo_user)
        response = self.client.get('/api/auth/me/')
        self.assertTrue(response.data['must_reset_password'])
        self.assertTrue(response.data['is_demo'])


class DemoUserCatalogVisibilityTests(BaseAPITestCase):
    """
    Phase 33: demo users see the FULL course catalog (every published course,
    regardless of Organization), with courses outside their own org's normal
    assignment flagged is_locked=True rather than hidden. Locked courses stay
    a teaser card only — retrieval and enrollment are still denied
    server-side, matching visible_courses_for_user exactly.
    """

    def setUp(self):
        super().setUp()
        self.demo_learner = User.objects.create_user(
            email='demo-catalog@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org,
            is_demo=True,
        )

    def _by_slug(self, response):
        return {c['slug']: c for c in response.data}

    def test_demo_user_sees_every_published_course_across_organizations(self):
        self.auth_as(self.demo_learner)
        response = self.client.get('/api/courses/')
        slugs = self._by_slug(response)
        self.assertIn(self.published_org_course.slug, slugs)
        self.assertIn(self.other_org_course.slug, slugs)
        self.assertIn(self.platform_course.slug, slugs)
        # Drafts are still withheld entirely — is_published gates content
        # readiness, not the demo teaser mechanism.
        self.assertNotIn(self.unpublished_org_course.slug, slugs)

    def test_own_org_course_is_unlocked_others_are_locked(self):
        self.auth_as(self.demo_learner)
        courses = self._by_slug(self.client.get('/api/courses/'))
        self.assertFalse(courses[self.published_org_course.slug]['is_locked'])
        self.assertTrue(courses[self.other_org_course.slug]['is_locked'])
        self.assertTrue(courses[self.platform_course.slug]['is_locked'])

    def test_unlocked_courses_sort_before_locked_ones(self):
        # platform_course and other_org_course are both locked for this demo
        # user and created after published_org_course (Course.Meta.ordering
        # is -created_at) — without the lock-status sort, they'd outrank the
        # one unlocked course. Assert the unlocked/locked split wins first.
        self.auth_as(self.demo_learner)
        response = self.client.get('/api/courses/')
        lock_flags = [row['is_locked'] for row in response.data]
        self.assertEqual(lock_flags, sorted(lock_flags))
        self.assertFalse(response.data[0]['is_locked'])
        self.assertEqual(response.data[0]['slug'], self.published_org_course.slug)

    def test_granted_platform_course_becomes_unlocked_for_demo_user(self):
        CourseAccess.objects.create(course=self.platform_course, organization=self.org)
        self.auth_as(self.demo_learner)
        courses = self._by_slug(self.client.get('/api/courses/'))
        self.assertFalse(courses[self.platform_course.slug]['is_locked'])

    def test_non_demo_learner_never_sees_is_locked_true(self):
        self.auth_as(self.learner)
        courses = self._by_slug(self.client.get('/api/courses/'))
        self.assertEqual(set(courses), {self.published_org_course.slug})
        self.assertFalse(courses[self.published_org_course.slug]['is_locked'])

    def test_force_navigating_to_a_locked_course_url_is_denied(self):
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/courses/{self.other_org_course.slug}/')
        self.assertEqual(response.status_code, 404)

    def test_demo_user_cannot_enroll_in_a_locked_course(self):
        self.auth_as(self.demo_learner)
        response = self.client.post('/api/enrollments/', {'course': self.other_org_course.id})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            Enrollment.objects.filter(user=self.demo_learner, course=self.other_org_course).exists()
        )

    def test_demo_user_can_retrieve_and_enroll_in_an_assigned_course(self):
        self.auth_as(self.demo_learner)
        retrieve = self.client.get(f'/api/courses/{self.published_org_course.slug}/')
        self.assertEqual(retrieve.status_code, 200)

        enroll = self.client.post('/api/enrollments/', {'course': self.published_org_course.id})
        self.assertEqual(enroll.status_code, 201)


class DemoLessonAccessTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.demo_learner = User.objects.create_user(
            email='demo@example.com', password='pass12345',
            role=User.Role.LEARNER, organization=self.org,
            is_demo=True,
        )
        Enrollment.objects.create(
            user=self.demo_learner, course=self.published_org_course, status=Enrollment.Status.IN_PROGRESS
        )
        Enrollment.objects.create(
            user=self.learner, course=self.published_org_course, status=Enrollment.Status.IN_PROGRESS
        )

        self.published_org_course.is_demo_available = True
        self.published_org_course.save()
        # lesson1 is granted; lesson2 is intentionally left locked.
        DemoLessonAccess.objects.create(course=self.published_org_course, lesson=self.lesson1)

    # --- Course detail tree ---

    def _lessons_by_id(self, response):
        return {lesson['id']: lesson for module in response.data['modules'] for lesson in module['lessons']}

    def test_demo_user_sees_locked_lesson_with_no_slides(self):
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/courses/{self.published_org_course.slug}/')
        lessons = self._lessons_by_id(response)
        self.assertFalse(lessons[self.lesson1.id]['is_locked'])
        self.assertTrue(lessons[self.lesson2.id]['is_locked'])
        self.assertEqual(lessons[self.lesson2.id]['slides'], [])

    def test_non_demo_user_sees_every_lesson_unlocked(self):
        self.auth_as(self.learner)
        response = self.client.get(f'/api/courses/{self.published_org_course.slug}/')
        lessons = self._lessons_by_id(response)
        self.assertFalse(lessons[self.lesson1.id]['is_locked'])
        self.assertFalse(lessons[self.lesson2.id]['is_locked'])
        self.assertGreater(len(lessons[self.lesson1.id]['slides']), 0)

    def test_demo_user_unaffected_when_course_is_not_demo_available(self):
        self.published_org_course.is_demo_available = False
        self.published_org_course.save()
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/courses/{self.published_org_course.slug}/')
        lessons = self._lessons_by_id(response)
        self.assertFalse(lessons[self.lesson2.id]['is_locked'])

    # --- Enrollment progress-writing endpoints ---

    def test_demo_user_cannot_complete_locked_lesson(self):
        enrollment = Enrollment.objects.get(user=self.demo_learner, course=self.published_org_course)
        self.auth_as(self.demo_learner)
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson2.id})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(enrollment.lesson_progress.filter(lesson=self.lesson2).exists())

    def test_demo_user_can_complete_granted_lesson(self):
        enrollment = Enrollment.objects.get(user=self.demo_learner, course=self.published_org_course)
        self.auth_as(self.demo_learner)
        response = self.client.post(f'/api/enrollments/{enrollment.id}/complete-lesson/', {'lesson': self.lesson1.id})
        self.assertEqual(response.status_code, 200)

    def test_demo_user_cannot_write_slide_progress_on_locked_lesson(self):
        locked_slide = Slide.objects.create(lesson=self.lesson2, order=1, slide_type=Slide.SlideType.CONTENT)
        enrollment = Enrollment.objects.get(user=self.demo_learner, course=self.published_org_course)
        self.auth_as(self.demo_learner)
        response = self.client.post(
            f'/api/enrollments/{enrollment.id}/slide-progress/', {'slide': locked_slide.id, 'completed': True}
        )
        self.assertEqual(response.status_code, 400)

    def test_non_demo_user_can_write_slide_progress_on_any_lesson(self):
        open_slide = Slide.objects.create(lesson=self.lesson2, order=1, slide_type=Slide.SlideType.CONTENT)
        enrollment = Enrollment.objects.get(user=self.learner, course=self.published_org_course)
        self.auth_as(self.learner)
        response = self.client.post(
            f'/api/enrollments/{enrollment.id}/slide-progress/', {'slide': open_slide.id, 'completed': True}
        )
        self.assertEqual(response.status_code, 200)

    # --- CONTENT (Element) read path ---

    def test_demo_user_cannot_fetch_elements_of_locked_lesson(self):
        locked_slide = Slide.objects.create(lesson=self.lesson2, order=1, slide_type=Slide.SlideType.CONTENT)
        Element.objects.create(slide=locked_slide, order=1, element_type=Element.ElementType.TEXT, rich_text='secret')
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/elements/?slide={locked_slide.id}')
        self.assertEqual(response.data, [])

    def test_demo_user_can_fetch_elements_of_granted_lesson(self):
        granted_slide = Slide.objects.create(lesson=self.lesson1, order=2, slide_type=Slide.SlideType.CONTENT)
        Element.objects.create(slide=granted_slide, order=1, element_type=Element.ElementType.TEXT, rich_text='hello')
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/elements/?slide={granted_slide.id}')
        self.assertEqual(len(response.data), 1)

    # --- QUIZ read path ---

    def test_demo_user_cannot_fetch_quiz_of_locked_lesson(self):
        locked_slide = Slide.objects.create(lesson=self.lesson2, order=2, slide_type=Slide.SlideType.QUIZ)
        locked_quiz = Quiz.objects.create(slide=locked_slide, title='Locked quiz', pass_percentage=50)
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/quizzes/{locked_quiz.id}/')
        self.assertEqual(response.status_code, 404)

    def test_demo_user_can_fetch_quiz_of_granted_lesson(self):
        # self.quiz's slide (self.quiz_slide) is on lesson1, which is granted.
        self.auth_as(self.demo_learner)
        response = self.client.get(f'/api/quizzes/{self.quiz.id}/')
        self.assertEqual(response.status_code, 200)

    # --- ASSIGNMENT submission write path ---

    def test_demo_user_cannot_submit_assignment_on_locked_lesson(self):
        locked_slide = Slide.objects.create(lesson=self.lesson2, order=3, slide_type=Slide.SlideType.ASSIGNMENT)
        assignment = Assignment.objects.create(slide=locked_slide, instructions='Do the thing')
        self.auth_as(self.demo_learner)
        response = self.client.post(
            '/api/assignment-submissions/', {'assignment': assignment.id, 'text_response': 'my answer'}
        )
        self.assertEqual(response.status_code, 400)

    # --- SCENARIO attempt write path ---

    def test_demo_user_cannot_submit_scenario_attempt_on_locked_lesson(self):
        locked_slide = Slide.objects.create(lesson=self.lesson2, order=4, slide_type=Slide.SlideType.SCENARIO)
        start_node = ScenarioNode.objects.create(slide=locked_slide, node_key='start', is_start=True)
        ending_choice = ScenarioChoice.objects.create(node=start_node, choice_text='End it', next_node=None)

        self.auth_as(self.demo_learner)
        response = self.client.post(
            '/api/scenario-attempts/', {'slide': locked_slide.id, 'path_taken': [ending_choice.id]}
        )
        self.assertEqual(response.status_code, 400)

    # --- Admin demo-lesson-access management endpoints ---

    def test_admin_can_grant_and_revoke_demo_lesson_access(self):
        self.auth_as(self.instructor)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/demo-lesson-access/', {'lesson': self.lesson2.id}
        )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(
            DemoLessonAccess.objects.filter(course=self.published_org_course, lesson=self.lesson2).exists()
        )

        response = self.client.delete(
            f'/api/courses/{self.published_org_course.slug}/demo-lesson-access/revoke/',
            {'lesson': self.lesson2.id},
            format='json',
        )
        self.assertEqual(response.status_code, 204)
        self.assertFalse(
            DemoLessonAccess.objects.filter(course=self.published_org_course, lesson=self.lesson2).exists()
        )

    def test_learner_cannot_manage_demo_lesson_access(self):
        self.auth_as(self.learner)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/demo-lesson-access/', {'lesson': self.lesson2.id}
        )
        self.assertEqual(response.status_code, 403)

    def test_demo_lesson_access_grant_rejects_lesson_from_another_course(self):
        other_module = Module.objects.create(course=self.other_org_course, title='Other', order=1)
        foreign_lesson = Lesson.objects.create(module=other_module, title='Foreign', order=1)
        self.auth_as(self.platform_admin)
        response = self.client.post(
            f'/api/courses/{self.published_org_course.slug}/demo-lesson-access/', {'lesson': foreign_lesson.id}
        )
        self.assertEqual(response.status_code, 404)


class SlideNarrationFlowTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.content_slide = Slide.objects.create(
            lesson=self.lesson1, order=1, title='Intro Slide', slide_type=Slide.SlideType.CONTENT,
        )
        Element.objects.create(
            slide=self.content_slide, order=1, element_type=Element.ElementType.TEXT,
            rich_text='<p>Welcome to your AML training.</p>',
        )
        self.empty_slide = Slide.objects.create(
            lesson=self.lesson1, order=2, title='Empty Slide', slide_type=Slide.SlideType.CONTENT,
        )

    def test_learner_cannot_generate_narration(self):
        self.auth_as(self.learner)
        response = self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})
        self.assertEqual(response.status_code, 403)

    def test_instructor_and_org_admin_cannot_generate_narration(self):
        for user in (self.instructor, self.org_admin):
            self.auth_as(user)
            response = self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})
            self.assertEqual(response.status_code, 403)

    def test_generate_rejects_invalid_language(self):
        self.auth_as(self.platform_admin)
        response = self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'fr'})
        self.assertEqual(response.status_code, 400)

    def test_generate_fails_when_slide_has_no_narratable_text(self):
        self.auth_as(self.platform_admin)
        response = self.client.post('/api/slide-narrations/generate/', {'slide': self.empty_slide.id, 'language': 'en'})
        self.assertEqual(response.status_code, 400)

    @patch('narration.services._synthesize_speech')
    @patch('narration.services._generate_script')
    def test_platform_admin_can_generate_narration(self, mock_generate_script, mock_synthesize_speech):
        mock_generate_script.return_value = 'Welcome to this training module.'
        mock_synthesize_speech.return_value = (b'fake-mp3-bytes', 'en-US-JennyNeural')

        self.auth_as(self.platform_admin)
        response = self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['script_text'], 'Welcome to this training module.')
        self.assertEqual(response.data['voice_name'], 'en-US-JennyNeural')
        self.assertIsNotNone(response.data['audio_file'])

        narration = SlideNarration.objects.get(slide=self.content_slide, language='en')
        self.assertEqual(narration.generated_by, self.platform_admin)

    @patch('narration.services._synthesize_speech')
    @patch('narration.services._generate_script')
    def test_regenerating_one_language_leaves_the_other_untouched(self, mock_generate_script, mock_synthesize_speech):
        mock_generate_script.side_effect = ['English v1', 'Nepali script', 'English v2']
        mock_synthesize_speech.side_effect = [
            (b'audio-en-1', 'en-US-JennyNeural'),
            (b'audio-ne', 'ne-NP-HemkalaNeural'),
            (b'audio-en-2', 'en-US-JennyNeural'),
        ]
        self.auth_as(self.platform_admin)

        self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})
        self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'ne'})
        self.assertEqual(SlideNarration.objects.filter(slide=self.content_slide).count(), 2)

        response = self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})
        self.assertEqual(response.data['script_text'], 'English v2')
        self.assertEqual(SlideNarration.objects.filter(slide=self.content_slide).count(), 2)

        ne_narration = SlideNarration.objects.get(slide=self.content_slide, language='ne')
        self.assertEqual(ne_narration.script_text, 'Nepali script')

    @patch('narration.services._synthesize_speech')
    @patch('narration.services._generate_script')
    def test_learner_can_view_generated_narration_for_a_visible_course(self, mock_generate_script, mock_synthesize_speech):
        mock_generate_script.return_value = 'Script text'
        mock_synthesize_speech.return_value = (b'audio-bytes', 'en-US-JennyNeural')
        self.auth_as(self.platform_admin)
        self.client.post('/api/slide-narrations/generate/', {'slide': self.content_slide.id, 'language': 'en'})

        self.auth_as(self.learner)
        response = self.client.get(f'/api/slide-narrations/?slide={self.content_slide.id}')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]['language'], 'en')

    def test_learner_in_other_org_cannot_see_narration_for_a_course_they_cannot_view(self):
        self.auth_as(self.other_org_learner)
        response = self.client.get(f'/api/slide-narrations/?slide={self.content_slide.id}')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 0)


class UserPreferenceApiTests(BaseAPITestCase):
    def test_learner_can_set_preferred_narration_language(self):
        self.auth_as(self.learner)
        self.assertEqual(self.client.get('/api/auth/me/').data['preferred_narration_language'], 'en')

        response = self.client.patch('/api/auth/me/', {'preferred_narration_language': 'ne'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['preferred_narration_language'], 'ne')

        self.learner.refresh_from_db()
        self.assertEqual(self.learner.preferred_narration_language, 'ne')

    def test_patch_me_cannot_change_role(self):
        self.auth_as(self.learner)
        response = self.client.patch('/api/auth/me/', {'preferred_narration_language': 'ne', 'role': 'PLATFORM_ADMIN'})
        self.assertEqual(response.status_code, 200)

        self.learner.refresh_from_db()
        self.assertEqual(self.learner.role, User.Role.LEARNER)


class LevelAssessmentAttemptServiceTests(TestCase):
    """
    Model + service coverage for the standalone role-based assessment system
    (independent of Course/Slide) — no API surface yet, so these exercise
    start_level_assessment_attempt directly, mirroring
    DemoUserProvisioningServiceTests' pattern for a service with no endpoint.
    """

    def setUp(self):
        self.org = Organization.objects.create(name='Acme Bank', slug='acme-bank-level')
        self.user = User.objects.create_user(email='learner@example.com', password='pw', role=User.Role.LEARNER)
        self.level = AssessmentLevel.objects.create(
            organization=self.org,
            name=User.AssessmentLevel.OFFICER,
            pass_threshold=70,
            questions_per_attempt=3,
        )
        self.set_a = QuestionSet.objects.create(assessment_level=self.level, label='Set 1')
        self.set_b = QuestionSet.objects.create(assessment_level=self.level, label='Set 2')
        # Questions spread across both sets — the draw pools them together.
        for question_set in (self.set_a, self.set_b):
            for i in range(2):
                question = LevelQuestion.objects.create(
                    question_set=question_set,
                    question_text=f'Question {question_set.label}-{i}',
                    question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
                )
                LevelChoice.objects.create(question=question, choice_text='Correct', is_correct=True)
                LevelChoice.objects.create(question=question, choice_text='Wrong', is_correct=False)

    def test_draws_the_configured_number_of_questions_from_the_combined_pool(self):
        attempt = start_level_assessment_attempt(user=self.user, assessment_level=self.level)

        self.assertEqual(len(attempt.questions_drawn), 3)
        pool_ids = set(LevelQuestion.objects.filter(question_set__assessment_level=self.level).values_list('id', flat=True))
        self.assertTrue(set(attempt.questions_drawn).issubset(pool_ids))
        self.assertEqual(len(set(attempt.questions_drawn)), 3)  # no duplicates

    def test_second_attempt_blocked_while_one_is_in_progress(self):
        start_level_assessment_attempt(user=self.user, assessment_level=self.level)

        with self.assertRaises(LevelAssessmentError):
            start_level_assessment_attempt(user=self.user, assessment_level=self.level)

        self.assertEqual(LevelAssessmentAttempt.objects.filter(user=self.user, assessment_level=self.level).count(), 1)

    def test_retake_allowed_after_prior_attempt_is_submitted(self):
        first = start_level_assessment_attempt(user=self.user, assessment_level=self.level)
        first.submitted_at = timezone.now()
        first.score_percent = Decimal('33.33')
        first.passed = False
        first.save()

        second = start_level_assessment_attempt(user=self.user, assessment_level=self.level)

        self.assertNotEqual(first.id, second.id)
        self.assertEqual(LevelAssessmentAttempt.objects.filter(user=self.user, assessment_level=self.level).count(), 2)

    def test_rejects_when_pool_smaller_than_questions_per_attempt(self):
        self.level.questions_per_attempt = 999
        self.level.save()

        with self.assertRaises(LevelAssessmentError):
            start_level_assessment_attempt(user=self.user, assessment_level=self.level)

        self.assertEqual(LevelAssessmentAttempt.objects.count(), 0)

    def test_different_users_may_each_have_their_own_open_attempt(self):
        other_user = User.objects.create_user(email='other@example.com', password='pw', role=User.Role.LEARNER)

        start_level_assessment_attempt(user=self.user, assessment_level=self.level)
        other_attempt = start_level_assessment_attempt(user=other_user, assessment_level=self.level)

        self.assertIsNotNone(other_attempt.id)


class LevelQuestionImportApiTests(BaseAPITestCase):
    def setUp(self):
        super().setUp()
        self.level = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, questions_per_attempt=2,
        )
        self.other_org_level = AssessmentLevel.objects.create(
            organization=self.other_org, name=User.AssessmentLevel.OFFICER, questions_per_attempt=2,
        )

    def import_url(self, level):
        return f'/api/assessment-levels/{level.id}/import-questions/'

    def test_learner_cannot_import(self):
        upload = make_question_template_upload({'Set 1': []})
        self.auth_as(self.learner)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')
        self.assertEqual(response.status_code, 403)

    def test_org_admin_cannot_import_into_another_organizations_level(self):
        upload = make_question_template_upload({'Set 1': []})
        self.auth_as(self.org_admin)
        response = self.client.post(self.import_url(self.other_org_level), {'file': upload}, format='multipart')
        self.assertEqual(response.status_code, 404)

    def test_valid_single_choice_and_multiple_answer_rows_are_created(self):
        rows = [
            ('Set 1', 'Capital of France?', 'Single Choice', 'Paris', 'Rome', 'Berlin', 'Madrid', '',
             'A', 2, 'Paris is correct.', 'Well done', 'Try again'),
            ('Set 1', 'Which are primary colors?', 'Multiple Answer', 'Red', 'Green', 'Blue', 'Purple', 'Yellow',
             'A, C', 1, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['created']), 2)
        self.assertEqual(response.data['failed'], [])

        question_set = QuestionSet.objects.get(assessment_level=self.level, label='Set 1')
        self.assertEqual(question_set.questions.count(), 2)

        single_choice = question_set.questions.get(question_type=LevelQuestion.QuestionType.SINGLE_CHOICE)
        self.assertEqual(single_choice.marks, 2)
        self.assertEqual(single_choice.choices.count(), 4)  # Option E left blank
        self.assertEqual(single_choice.choices.get(is_correct=True).choice_text, 'Paris')

        multiple_answer = question_set.questions.get(question_type=LevelQuestion.QuestionType.MULTIPLE_ANSWER)
        self.assertEqual(multiple_answer.choices.count(), 5)
        self.assertEqual(
            set(multiple_answer.choices.filter(is_correct=True).values_list('choice_text', flat=True)),
            {'Red', 'Blue'},
        )

    def test_reuses_existing_question_set_by_label(self):
        existing_set = QuestionSet.objects.create(assessment_level=self.level, label='Set 1')
        rows = [
            ('Set 1', 'Q1?', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(QuestionSet.objects.filter(assessment_level=self.level, label='Set 1').count(), 1)
        self.assertEqual(existing_set.questions.count(), 1)

    def test_missing_options_and_bad_question_type_are_reported_not_dropped(self):
        rows = [
            # Missing Option D
            ('Set 1', 'Bad row 1', 'Single Choice', 'A', 'B', 'C', '', '', 'A', 1, '', '', ''),
            # Invalid Question Type
            ('Set 1', 'Bad row 2', 'Essay', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', ''),
            # Valid row in between — must still be created despite the failures around it
            ('Set 1', 'Good row', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['created']), 1)
        self.assertEqual(len(response.data['failed']), 2)
        self.assertIn('Option D', response.data['failed'][0]['reason'])
        self.assertIn('Single Choice', response.data['failed'][1]['reason'])
        self.assertEqual(response.data['failed'][0]['row'], 2)
        self.assertEqual(response.data['failed'][1]['row'], 3)

    def test_correct_answer_referencing_empty_option_is_rejected(self):
        rows = [
            ('Set 1', 'Bad row', 'Single Choice', 'A', 'B', 'C', 'D', '', 'E', 1, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertIn('empty option', response.data['failed'][0]['reason'])

    def test_single_choice_with_multiple_correct_answers_is_rejected(self):
        rows = [
            ('Set 1', 'Bad row', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A,B', 1, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertIn('exactly one', response.data['failed'][0]['reason'])

    def test_non_positive_marks_is_rejected(self):
        rows = [
            ('Set 1', 'Bad row', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 0, '', '', ''),
        ]
        upload = make_question_template_upload({'Sheet1': rows})

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['created'], [])
        self.assertIn('positive whole number', response.data['failed'][0]['reason'])

    def test_multiple_sheets_are_all_parsed(self):
        rows_by_sheet = {
            'Sheet A': [('Set 1', 'Q1?', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', '')],
            'Sheet B': [('Set 2', 'Q2?', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', '')],
        }
        upload = make_question_template_upload(rows_by_sheet)

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['created']), 2)
        self.assertEqual(
            QuestionSet.objects.filter(assessment_level=self.level).count(), 2,
        )

    def test_sheet_with_missing_required_column_is_reported_without_aborting_other_sheets(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        bad_sheet = workbook.create_sheet('Bad Sheet')
        bad_sheet.append(['Question Set', 'Question Text'])  # missing most required columns
        bad_sheet.append(['Set 1', 'Q1?'])
        good_sheet = workbook.create_sheet('Good Sheet')
        good_sheet.append(LEVEL_QUESTION_TEMPLATE_HEADER)
        good_sheet.append(('Set 2', 'Q2?', 'Single Choice', 'A', 'B', 'C', 'D', '', 'A', 1, '', '', ''))
        buffer = io.BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile('questions.xlsx', buffer.getvalue())

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data['created']), 1)
        self.assertEqual(len(response.data['failed']), 1)
        self.assertIsNone(response.data['failed'][0]['row'])
        self.assertIn('Missing required column', response.data['failed'][0]['reason'])

    def test_non_xlsx_upload_returns_400(self):
        upload = SimpleUploadedFile('questions.xlsx', b'not a real workbook', content_type='text/plain')

        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, 400)

    def test_import_requires_a_file(self):
        self.auth_as(self.instructor)
        response = self.client.post(self.import_url(self.level), {}, format='multipart')
        self.assertEqual(response.status_code, 400)


class LevelAssessmentStudentFlowApiTests(BaseAPITestCase):
    """
    The student-facing flow: dashboard/landing status lookup, starting an
    attempt (random draw, answer-key stripped), submitting it for grading
    (answer-key revealed per-answer only after submission), and the
    resulting status transitions (Not started -> In progress -> Passed/Failed
    -> retake available).
    """

    def setUp(self):
        super().setUp()
        self.level = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, pass_threshold=50, questions_per_attempt=2,
        )
        question_set = QuestionSet.objects.create(assessment_level=self.level, label='Set 1')

        self.q1 = LevelQuestion.objects.create(
            question_set=question_set, question_text='Q1?', question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
            marks=1,
        )
        self.q1_correct = LevelChoice.objects.create(question=self.q1, choice_text='A', is_correct=True)
        LevelChoice.objects.create(question=self.q1, choice_text='B', is_correct=False)

        self.q2 = LevelQuestion.objects.create(
            question_set=question_set, question_text='Q2?', question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
            marks=1,
        )
        self.q2_correct = LevelChoice.objects.create(question=self.q2, choice_text='A', is_correct=True)
        LevelChoice.objects.create(question=self.q2, choice_text='B', is_correct=False)

        self.q3 = LevelQuestion.objects.create(
            question_set=question_set, question_text='Q3?', question_type=LevelQuestion.QuestionType.MULTIPLE_ANSWER,
            marks=2,
        )
        self.q3_correct_a = LevelChoice.objects.create(question=self.q3, choice_text='A', is_correct=True)
        self.q3_correct_b = LevelChoice.objects.create(question=self.q3, choice_text='B', is_correct=True)
        LevelChoice.objects.create(question=self.q3, choice_text='C', is_correct=False)

        self.learner.assessment_level = User.AssessmentLevel.OFFICER
        self.learner.save()

    def correct_choice_ids_for(self, question_id):
        return list(LevelQuestion.objects.get(id=question_id).choices.filter(is_correct=True).values_list('id', flat=True))

    def start_attempt(self):
        self.auth_as(self.learner)
        return self.client.post('/api/level-attempts/start/')

    def submit_all_correct(self, attempt_id, questions):
        answers = [{'question': q['id'], 'selected_choices': self.correct_choice_ids_for(q['id'])} for q in questions]
        return self.client.post(f'/api/level-attempts/{attempt_id}/submit/', {'answers': answers}, format='json')

    def test_not_assigned_when_user_has_no_assessment_level(self):
        self.learner.assessment_level = None
        self.learner.save()
        self.auth_as(self.learner)

        response = self.client.get('/api/my-assessment-level/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, {'assigned': False})

    def test_not_assigned_when_organization_has_no_configured_level(self):
        self.other_org_learner.assessment_level = User.AssessmentLevel.OFFICER
        self.other_org_learner.save()  # other_org has no AssessmentLevel configured
        self.auth_as(self.other_org_learner)

        response = self.client.get('/api/my-assessment-level/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, {'assigned': False})

    def test_status_not_started_before_any_attempt(self):
        self.auth_as(self.learner)
        response = self.client.get('/api/my-assessment-level/')

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['assigned'])
        self.assertEqual(response.data['status'], 'NOT_STARTED')
        self.assertIsNone(response.data['open_attempt_id'])
        self.assertEqual(response.data['assessment_level']['questions_per_attempt'], 2)

    def test_start_draws_the_configured_number_with_no_answer_key_exposed(self):
        response = self.start_attempt()

        self.assertEqual(response.status_code, 201)
        questions = response.data['questions']
        self.assertEqual(len(questions), 2)
        for question in questions:
            for choice in question['choices']:
                self.assertNotIn('is_correct', choice)
        self.assertEqual(response.data['answers'], [])

    def test_status_in_progress_after_starting_and_blocks_a_second_start(self):
        self.start_attempt()

        status_response = self.client.get('/api/my-assessment-level/')
        self.assertEqual(status_response.data['status'], 'IN_PROGRESS')
        self.assertIsNotNone(status_response.data['open_attempt_id'])

        second_start = self.client.post('/api/level-attempts/start/')
        self.assertEqual(second_start.status_code, 400)

    def test_submit_all_correct_passes_and_reveals_answer_key(self):
        start_response = self.start_attempt()
        attempt_id = start_response.data['id']
        questions = start_response.data['questions']

        response = self.submit_all_correct(attempt_id, questions)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['passed'])
        self.assertEqual(Decimal(response.data['score_percent']), Decimal('100.00'))
        for answer in response.data['answers']:
            self.assertTrue(answer['is_correct'])
            self.assertEqual(set(answer['correct_choice_ids']), set(answer['selected_choices']))

        status_response = self.client.get('/api/my-assessment-level/')
        self.assertEqual(status_response.data['status'], 'PASSED')
        self.assertIsNone(status_response.data['open_attempt_id'])

    def test_submit_all_wrong_fails(self):
        start_response = self.start_attempt()
        attempt_id = start_response.data['id']
        questions = start_response.data['questions']

        answers = []
        for question in questions:
            wrong_choice = next(
                c['id'] for c in question['choices'] if c['id'] not in self.correct_choice_ids_for(question['id'])
            )
            answers.append({'question': question['id'], 'selected_choices': [wrong_choice]})
        response = self.client.post(f'/api/level-attempts/{attempt_id}/submit/', {'answers': answers}, format='json')

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.data['passed'])

        status_response = self.client.get('/api/my-assessment-level/')
        self.assertEqual(status_response.data['status'], 'FAILED')

    def test_retake_allowed_after_submission_with_fresh_draw(self):
        start_response = self.start_attempt()
        self.submit_all_correct(start_response.data['id'], start_response.data['questions'])

        second_start = self.client.post('/api/level-attempts/start/')
        self.assertEqual(second_start.status_code, 201)
        self.assertNotEqual(second_start.data['id'], start_response.data['id'])

    def test_cannot_submit_answers_missing_a_drawn_question(self):
        start_response = self.start_attempt()
        attempt_id = start_response.data['id']
        questions = start_response.data['questions']

        answers = [{'question': questions[0]['id'], 'selected_choices': self.correct_choice_ids_for(questions[0]['id'])}]
        response = self.client.post(f'/api/level-attempts/{attempt_id}/submit/', {'answers': answers}, format='json')

        self.assertEqual(response.status_code, 400)

    def test_cannot_submit_a_question_not_drawn_for_this_attempt(self):
        start_response = self.start_attempt()
        attempt_id = start_response.data['id']
        questions = start_response.data['questions']
        drawn_ids = {q['id'] for q in questions}
        not_drawn = next(q for q in (self.q1, self.q2, self.q3) if q.id not in drawn_ids)

        answers = [{'question': q['id'], 'selected_choices': self.correct_choice_ids_for(q['id'])} for q in questions]
        answers[0]['question'] = not_drawn.id
        response = self.client.post(f'/api/level-attempts/{attempt_id}/submit/', {'answers': answers}, format='json')

        self.assertEqual(response.status_code, 400)

    def test_cannot_submit_the_same_attempt_twice(self):
        start_response = self.start_attempt()
        self.submit_all_correct(start_response.data['id'], start_response.data['questions'])

        second_submit = self.client.post(f'/api/level-attempts/{start_response.data["id"]}/submit/', {'answers': []}, format='json')
        self.assertEqual(second_submit.status_code, 400)

    def test_another_user_cannot_view_or_submit_someone_elses_attempt(self):
        start_response = self.start_attempt()
        attempt_id = start_response.data['id']

        self.other_org_learner.organization = self.org
        self.other_org_learner.assessment_level = User.AssessmentLevel.OFFICER
        self.other_org_learner.save()
        self.auth_as(self.other_org_learner)

        get_response = self.client.get(f'/api/level-attempts/{attempt_id}/')
        self.assertEqual(get_response.status_code, 404)

        submit_response = self.client.post(f'/api/level-attempts/{attempt_id}/submit/', {'answers': []}, format='json')
        self.assertEqual(submit_response.status_code, 404)

    def test_start_requires_an_assigned_assessment_level(self):
        self.learner.assessment_level = None
        self.learner.save()
        response = self.start_attempt()
        self.assertEqual(response.status_code, 400)


class LevelAssessmentBadgeTests(TestCase):
    """
    gamification.services.award_badges_for_level_assessment_attempt — the
    five level-assessment-specific achievement conditions, exercised directly
    against manually-built attempts/answers (rather than the random-draw
    service) so each scenario's exact sequence/history is under test control.
    """

    def setUp(self):
        self.org = Organization.objects.create(name='Acme Bank', slug='acme-bank-badges')
        self.level = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, pass_threshold=50, questions_per_attempt=5,
        )
        question_set = QuestionSet.objects.create(assessment_level=self.level, label='Set 1')
        self.questions = []
        for i in range(5):
            question = LevelQuestion.objects.create(
                question_set=question_set, question_text=f'Q{i}?',
                question_type=LevelQuestion.QuestionType.SINGLE_CHOICE, marks=1,
            )
            LevelChoice.objects.create(question=question, choice_text='Correct', is_correct=True)
            LevelChoice.objects.create(question=question, choice_text='Wrong', is_correct=False)
            self.questions.append(question)

    def make_user(self, email, branch_department=''):
        return User.objects.create_user(
            email=email, password='pw', role=User.Role.LEARNER, organization=self.org,
            assessment_level=User.AssessmentLevel.OFFICER, branch_department=branch_department,
        )

    def make_attempt(self, user, correctness, passed, questions_drawn=None):
        """`correctness` is a list of bools in drawn order, one per question in self.questions[:len(correctness)]."""
        questions = questions_drawn or self.questions[: len(correctness)]
        attempt = LevelAssessmentAttempt.objects.create(
            user=user, assessment_level=self.level, questions_drawn=[q.id for q in questions],
            submitted_at=timezone.now(), passed=passed,
            score_percent=Decimal('100.00') if all(correctness) else Decimal('40.00'),
        )
        for question, is_correct in zip(questions, correctness):
            LevelAssessmentAnswer.objects.create(attempt=attempt, question=question, is_correct=is_correct)
        return attempt

    def earned_keys(self, user):
        return set(UserBadge.objects.filter(user=user).values_list('badge__key', flat=True))

    # --- First Strike ---

    def test_first_strike_awarded_on_first_attempt_with_a_correct_answer(self):
        user = self.make_user('a@example.com')
        attempt = self.make_attempt(user, [True, False], passed=False)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertIn('first_strike', self.earned_keys(user))

    def test_first_strike_not_awarded_if_first_attempt_all_wrong(self):
        user = self.make_user('b@example.com')
        attempt = self.make_attempt(user, [False, False], passed=False)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertNotIn('first_strike', self.earned_keys(user))

    def test_first_strike_not_awarded_on_a_later_attempt(self):
        user = self.make_user('c@example.com')
        self.make_attempt(user, [False, False], passed=False)  # first attempt, all wrong
        second_attempt = self.make_attempt(user, [True, True], passed=False)  # second attempt, correct

        award_badges_for_level_assessment_attempt(second_attempt)

        self.assertNotIn('first_strike', self.earned_keys(user))

    # --- Hat Trick ---

    def test_hat_trick_awarded_for_three_consecutive_correct_in_drawn_order(self):
        user = self.make_user('d@example.com')
        attempt = self.make_attempt(user, [True, True, True, False, False], passed=False)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertIn('hat_trick', self.earned_keys(user))

    def test_hat_trick_awarded_for_a_trailing_streak(self):
        user = self.make_user('e@example.com')
        attempt = self.make_attempt(user, [True, False, True, True, True], passed=False)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertIn('hat_trick', self.earned_keys(user))

    def test_hat_trick_not_awarded_without_three_in_a_row(self):
        user = self.make_user('f@example.com')
        attempt = self.make_attempt(user, [True, True, False, True, True], passed=False)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertNotIn('hat_trick', self.earned_keys(user))

    # --- Perfect Score ---

    def test_perfect_score_awarded_for_100_percent_level_assessment(self):
        user = self.make_user('g@example.com')
        attempt = self.make_attempt(user, [True] * 5, passed=True)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertIn('perfect_score', self.earned_keys(user))

    def test_perfect_score_not_awarded_below_100_percent(self):
        user = self.make_user('h@example.com')
        attempt = self.make_attempt(user, [True, True, True, True, False], passed=True)
        attempt.score_percent = Decimal('80.00')
        attempt.save()

        award_badges_for_level_assessment_attempt(attempt)

        self.assertNotIn('perfect_score', self.earned_keys(user))

    # --- Comeback ---

    def test_comeback_awarded_after_a_prior_failed_attempt_at_the_same_level(self):
        user = self.make_user('i@example.com')
        self.make_attempt(user, [False, False], passed=False)
        second_attempt = self.make_attempt(user, [True, True], passed=True)

        award_badges_for_level_assessment_attempt(second_attempt)

        self.assertIn('comeback', self.earned_keys(user))

    def test_comeback_not_awarded_when_first_attempt_already_passed(self):
        user = self.make_user('j@example.com')
        attempt = self.make_attempt(user, [True, True], passed=True)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertNotIn('comeback', self.earned_keys(user))

    def test_comeback_not_awarded_when_the_retake_also_fails(self):
        user = self.make_user('k@example.com')
        self.make_attempt(user, [False, False], passed=False)
        second_attempt = self.make_attempt(user, [False, False], passed=False)

        award_badges_for_level_assessment_attempt(second_attempt)

        self.assertNotIn('comeback', self.earned_keys(user))

    # --- Branch Pride ---

    def test_branch_pride_awarded_to_first_passer_in_a_branch(self):
        user = self.make_user('l@example.com', branch_department='Head Office')
        attempt = self.make_attempt(user, [True, True], passed=True)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertIn('branch_pride', self.earned_keys(user))

    def test_branch_pride_not_awarded_to_second_passer_in_the_same_branch(self):
        first_user = self.make_user('m@example.com', branch_department='Head Office')
        first_attempt = self.make_attempt(first_user, [True, True], passed=True)
        award_badges_for_level_assessment_attempt(first_attempt)

        second_user = self.make_user('n@example.com', branch_department='Head Office')
        second_attempt = self.make_attempt(second_user, [True, True], passed=True)
        award_badges_for_level_assessment_attempt(second_attempt)

        self.assertNotIn('branch_pride', self.earned_keys(second_user))

    def test_branch_pride_awarded_independently_per_branch(self):
        first_user = self.make_user('o@example.com', branch_department='Head Office')
        award_badges_for_level_assessment_attempt(self.make_attempt(first_user, [True, True], passed=True))

        other_branch_user = self.make_user('p@example.com', branch_department='Pokhara Branch')
        other_attempt = self.make_attempt(other_branch_user, [True, True], passed=True)
        award_badges_for_level_assessment_attempt(other_attempt)

        self.assertIn('branch_pride', self.earned_keys(other_branch_user))

    def test_branch_pride_not_awarded_without_a_branch_department(self):
        user = self.make_user('q@example.com', branch_department='')
        attempt = self.make_attempt(user, [True, True], passed=True)

        award_badges_for_level_assessment_attempt(attempt)

        self.assertNotIn('branch_pride', self.earned_keys(user))

    # --- Idempotency / already-earned ---

    def test_badge_not_re_awarded_if_already_earned(self):
        user = self.make_user('r@example.com')
        badge = Badge.objects.get(key='first_strike')
        UserBadge.objects.create(user=user, badge=badge)

        attempt = self.make_attempt(user, [True, True], passed=False)
        award_badges_for_level_assessment_attempt(attempt)

        self.assertEqual(UserBadge.objects.filter(user=user, badge=badge).count(), 1)


class LevelAssessmentBadgeIntegrationTests(BaseAPITestCase):
    """Confirms the submit endpoint actually wires up badge awarding, not just the service function in isolation."""

    def test_submitting_a_perfect_first_attempt_awards_first_strike_and_perfect_score(self):
        level = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, pass_threshold=50, questions_per_attempt=1,
        )
        question_set = QuestionSet.objects.create(assessment_level=level, label='Set 1')
        question = LevelQuestion.objects.create(
            question_set=question_set, question_text='Q1?', question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
            marks=1,
        )
        correct_choice = LevelChoice.objects.create(question=question, choice_text='Correct', is_correct=True)
        LevelChoice.objects.create(question=question, choice_text='Wrong', is_correct=False)

        self.learner.assessment_level = User.AssessmentLevel.OFFICER
        self.learner.save()
        self.auth_as(self.learner)

        start_response = self.client.post('/api/level-attempts/start/')
        attempt_id = start_response.data['id']

        submit_response = self.client.post(
            f'/api/level-attempts/{attempt_id}/submit/',
            {'answers': [{'question': question.id, 'selected_choices': [correct_choice.id]}]},
            format='json',
        )

        self.assertEqual(submit_response.status_code, 200)
        earned = set(UserBadge.objects.filter(user=self.learner).values_list('badge__key', flat=True))
        self.assertIn('first_strike', earned)
        self.assertIn('perfect_score', earned)


class LeaderboardLevelAssessmentPointsTests(TestCase):
    """
    gamification.services.recalculate_leaderboard_entry's level-assessment
    contribution — exercised directly against manually-built attempts so the
    exact pass/fail/retake history is under test control.
    """

    def setUp(self):
        self.org = Organization.objects.create(name='Acme Bank', slug='acme-bank-points')
        self.user = User.objects.create_user(
            email='points@example.com', password='pw', role=User.Role.LEARNER, organization=self.org,
        )
        self.level_a = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, questions_per_attempt=1,
        )
        self.level_b = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.MANAGEMENT, questions_per_attempt=1,
        )

    def make_attempt(self, level, passed):
        return LevelAssessmentAttempt.objects.create(
            user=self.user, assessment_level=level, questions_drawn=[], submitted_at=timezone.now(), passed=passed,
        )

    def test_a_passed_level_assessment_is_worth_more_than_a_single_course_completion(self):
        # The task's explicit requirement: weighted above, not equal to, an
        # ordinary course completion, reflecting its higher-stakes status.
        self.assertGreater(LEVEL_ASSESSMENT_PASS_POINTS, COURSE_COMPLETION_POINTS)

    def test_passed_level_assessment_awards_configured_points(self):
        self.make_attempt(self.level_a, passed=True)

        entry = recalculate_leaderboard_entry(self.user)

        self.assertEqual(entry.total_points, LEVEL_ASSESSMENT_PASS_POINTS)
        self.assertEqual(entry.level_assessments_passed_count, 1)

    def test_failed_attempt_awards_no_points(self):
        self.make_attempt(self.level_a, passed=False)

        entry = recalculate_leaderboard_entry(self.user)

        self.assertEqual(entry.total_points, 0)
        self.assertEqual(entry.level_assessments_passed_count, 0)

    def test_retaking_an_already_passed_level_does_not_double_count(self):
        self.make_attempt(self.level_a, passed=False)
        self.make_attempt(self.level_a, passed=True)
        self.make_attempt(self.level_a, passed=True)  # retake after already having passed

        entry = recalculate_leaderboard_entry(self.user)

        self.assertEqual(entry.total_points, LEVEL_ASSESSMENT_PASS_POINTS)
        self.assertEqual(entry.level_assessments_passed_count, 1)

    def test_passing_two_distinct_levels_counts_both(self):
        self.make_attempt(self.level_a, passed=True)
        self.make_attempt(self.level_b, passed=True)

        entry = recalculate_leaderboard_entry(self.user)

        self.assertEqual(entry.total_points, LEVEL_ASSESSMENT_PASS_POINTS * 2)
        self.assertEqual(entry.level_assessments_passed_count, 2)

    def test_combines_with_course_completion_points(self):
        course = Course.objects.create(
            title='Compliance 101', slug='compliance-101-points', organization=self.org,
            content_owner=Course.ContentOwner.ORGANIZATION,
        )
        Enrollment.objects.create(user=self.user, course=course, status=Enrollment.Status.COMPLETED)
        self.make_attempt(self.level_a, passed=True)

        entry = recalculate_leaderboard_entry(self.user)

        self.assertEqual(entry.total_points, COURSE_COMPLETION_POINTS + LEVEL_ASSESSMENT_PASS_POINTS)


class LeaderboardLevelAssessmentIntegrationTests(BaseAPITestCase):
    """Confirms level assessment points flow through the real submit endpoint, and that the
    leaderboard stays organization-scoped (Phase 25's original hard requirement) once they do."""

    def setUp(self):
        super().setUp()
        self.level = AssessmentLevel.objects.create(
            organization=self.org, name=User.AssessmentLevel.OFFICER, pass_threshold=50, questions_per_attempt=1,
        )
        question_set = QuestionSet.objects.create(assessment_level=self.level, label='Set 1')
        self.question = LevelQuestion.objects.create(
            question_set=question_set, question_text='Q1?', question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
            marks=1,
        )
        self.correct_choice = LevelChoice.objects.create(question=self.question, choice_text='Correct', is_correct=True)
        LevelChoice.objects.create(question=self.question, choice_text='Wrong', is_correct=False)

        self.learner.assessment_level = User.AssessmentLevel.OFFICER
        self.learner.save()

    def test_passing_updates_leaderboard_points_through_the_submit_endpoint(self):
        self.auth_as(self.learner)
        start_response = self.client.post('/api/level-attempts/start/')
        attempt_id = start_response.data['id']

        submit_response = self.client.post(
            f'/api/level-attempts/{attempt_id}/submit/',
            {'answers': [{'question': self.question.id, 'selected_choices': [self.correct_choice.id]}]},
            format='json',
        )

        self.assertEqual(submit_response.status_code, 200)
        entry = LeaderboardEntry.objects.get(user=self.learner)
        self.assertEqual(entry.total_points, LEVEL_ASSESSMENT_PASS_POINTS)
        self.assertEqual(entry.level_assessments_passed_count, 1)

    def test_leaderboard_endpoint_stays_organization_scoped_after_a_level_assessment_pass(self):
        # A learner in a *different* organization passes their own level
        # assessment — its points must never surface in self.org's leaderboard.
        other_level = AssessmentLevel.objects.create(
            organization=self.other_org, name=User.AssessmentLevel.OFFICER, pass_threshold=50, questions_per_attempt=1,
        )
        other_question_set = QuestionSet.objects.create(assessment_level=other_level, label='Set 1')
        other_question = LevelQuestion.objects.create(
            question_set=other_question_set, question_text='Q?', question_type=LevelQuestion.QuestionType.SINGLE_CHOICE,
            marks=1,
        )
        other_correct = LevelChoice.objects.create(question=other_question, choice_text='Correct', is_correct=True)
        LevelChoice.objects.create(question=other_question, choice_text='Wrong', is_correct=False)

        self.other_org_learner.assessment_level = User.AssessmentLevel.OFFICER
        self.other_org_learner.save()
        self.auth_as(self.other_org_learner)
        start_response = self.client.post('/api/level-attempts/start/')
        self.client.post(
            f'/api/level-attempts/{start_response.data["id"]}/submit/',
            {'answers': [{'question': other_question.id, 'selected_choices': [other_correct.id]}]},
            format='json',
        )

        self.auth_as(self.learner)
        response = self.client.get('/api/leaderboard/')

        self.assertEqual(response.status_code, 200)
        self.assertNotIn(self.other_org_learner.id, [row['user_id'] for row in response.data])
        # And confirm the other org's entry really was created with points —
        # this is a scoping check, not a "nothing happened" false negative.
        other_entry = LeaderboardEntry.objects.get(user=self.other_org_learner)
        self.assertEqual(other_entry.total_points, LEVEL_ASSESSMENT_PASS_POINTS)
        self.assertEqual(other_entry.organization_id, self.other_org.id)
