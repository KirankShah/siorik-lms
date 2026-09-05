"""
Admin-only bulk import of LevelQuestion/LevelChoice records from the "Level
Assessment Question Template" spreadsheet, for a specific AssessmentLevel.
Reporting follows the same pattern as accounts.views.DemoUserViewSet.bulk's
CSV upload: each row is validated and, if valid, committed immediately;
invalid rows are never silently dropped — they're collected with a reason
and reported back alongside whatever did succeed.
"""
from openpyxl import load_workbook

from .models import LevelChoice, LevelQuestion, QuestionSet

OPTION_LETTERS = ['A', 'B', 'C', 'D', 'E']
REQUIRED_OPTION_LETTERS = ['A', 'B', 'C', 'D']

# Normalized (lowercased, trimmed) header cell -> internal field name. Column
# order in the uploaded sheet doesn't matter, only that every one of these is
# present somewhere in the header row.
TEMPLATE_COLUMNS = {
    'question set': 'question_set',
    'question text': 'question_text',
    'question type': 'question_type',
    'option a': 'option_a',
    'option b': 'option_b',
    'option c': 'option_c',
    'option d': 'option_d',
    'option e': 'option_e',
    'correct answer(s)': 'correct_answers',
    'marks': 'marks',
    'explanation': 'explanation',
    'feedback if correct': 'feedback_correct',
    'feedback if incorrect': 'feedback_incorrect',
}

QUESTION_TYPE_LABELS = {
    'single choice': LevelQuestion.QuestionType.SINGLE_CHOICE,
    'multiple answer': LevelQuestion.QuestionType.MULTIPLE_ANSWER,
}


class LevelQuestionImportError(Exception):
    """Raised only for a whole-file failure (the upload can't be read as an
    .xlsx workbook at all) — there's no single row to blame, so this isn't
    reported through the per-row failure list."""


def _cell_text(value):
    return '' if value is None else str(value).strip()


def _column_index_map(header_row):
    """Maps internal field name -> column index from a header row. Raises
    ValueError (caller turns this into one per-sheet failure entry, not a
    whole-file error) listing whichever required column headers are absent."""
    index_by_field = {}
    for index, cell in enumerate(header_row):
        field = TEMPLATE_COLUMNS.get(_cell_text(cell).lower())
        if field:
            index_by_field[field] = index

    missing_fields = set(TEMPLATE_COLUMNS.values()) - set(index_by_field)
    if missing_fields:
        missing_labels = sorted(label for label, field in TEMPLATE_COLUMNS.items() if field in missing_fields)
        raise ValueError(f'Missing required column(s): {", ".join(missing_labels)}.')
    return index_by_field


def _parse_correct_answers(raw, question_type):
    letters = [part.strip().upper() for part in raw.split(',') if part.strip()]
    if not letters:
        return None, 'Correct Answer(s) is required.'
    invalid_letters = [letter for letter in letters if letter not in OPTION_LETTERS]
    if invalid_letters:
        return None, f'Correct Answer(s) "{raw}" must reference option letters A-E only.'
    if len(set(letters)) != len(letters):
        return None, f'Correct Answer(s) "{raw}" lists the same option more than once.'
    if question_type == LevelQuestion.QuestionType.SINGLE_CHOICE and len(letters) != 1:
        return None, 'Single Choice questions must have exactly one Correct Answer.'
    return letters, None


def _parse_marks(raw):
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return None, f'Marks "{raw}" is not a number.'
    if value <= 0 or value != int(value):
        return None, f'Marks "{raw}" must be a positive whole number.'
    return int(value), None


def _parse_row(row, column_index):
    def col(field):
        index = column_index[field]
        return _cell_text(row[index]) if index < len(row) else ''

    question_set_label = col('question_set')
    question_text = col('question_text')
    question_type_raw = col('question_type')
    options = {letter: col(f'option_{letter.lower()}') for letter in OPTION_LETTERS}
    correct_answers_raw = col('correct_answers')
    marks_raw = col('marks')

    if not question_set_label:
        return None, 'Missing Question Set.'
    if not question_text:
        return None, 'Missing Question Text.'

    missing_options = [letter for letter in REQUIRED_OPTION_LETTERS if not options[letter]]
    if missing_options:
        return None, f'Option {", ".join(missing_options)} must be filled in.'

    question_type = QUESTION_TYPE_LABELS.get(question_type_raw.lower())
    if question_type is None:
        return None, f'Question Type "{question_type_raw}" must be "Single Choice" or "Multiple Answer".'

    letters, error = _parse_correct_answers(correct_answers_raw, question_type)
    if error:
        return None, error

    missing_referenced = [letter for letter in letters if not options[letter]]
    if missing_referenced:
        return None, f'Correct Answer(s) references empty option(s): {", ".join(missing_referenced)}.'

    marks, error = _parse_marks(marks_raw)
    if error:
        return None, error

    return {
        'question_set_label': question_set_label,
        'question_text': question_text,
        'question_type': question_type,
        'options': options,
        'correct_letters': letters,
        'marks': marks,
        'explanation': col('explanation'),
        'feedback_correct': col('feedback_correct'),
        'feedback_incorrect': col('feedback_incorrect'),
    }, None


def _create_question(assessment_level, parsed):
    question_set, _ = QuestionSet.objects.get_or_create(
        assessment_level=assessment_level, label=parsed['question_set_label']
    )
    question = LevelQuestion.objects.create(
        question_set=question_set,
        question_text=parsed['question_text'],
        question_type=parsed['question_type'],
        marks=parsed['marks'],
        explanation=parsed['explanation'],
        feedback_correct=parsed['feedback_correct'],
        feedback_incorrect=parsed['feedback_incorrect'],
    )
    for order, letter in enumerate(OPTION_LETTERS):
        option_text = parsed['options'][letter]
        if not option_text:
            continue
        LevelChoice.objects.create(
            question=question,
            choice_text=option_text,
            is_correct=letter in parsed['correct_letters'],
            order=order,
        )
    return question, question_set


def import_level_questions(*, assessment_level, workbook_file):
    """
    Parses every sheet of the uploaded Level Assessment Question Template
    workbook, creating a QuestionSet (get-or-created by each row's "Question
    Set" column label, under `assessment_level`) plus LevelQuestion/
    LevelChoice rows for every row that validates.

    Each row is validated then, if valid, committed immediately — same
    pattern as accounts.views.DemoUserViewSet.bulk's CSV upload. A row that
    fails validation is skipped and reported in `failed` with its sheet name,
    row number, and reason; it never silently disappears, and it never
    blocks any other row (including other rows in the same Question Set) from
    being imported. A sheet whose header row is missing a required template
    column is reported the same way, as a single failure for that sheet, and
    the rest of the workbook is still processed.

    Raises LevelQuestionImportError only when the upload itself can't be
    read as an .xlsx workbook — there's no row or sheet to attribute that to.

    Returns (created, failed):
      created -- list of {'sheet', 'row', 'question_set'} for each row committed.
      failed  -- list of {'sheet', 'row', 'reason'} for each row (or sheet) rejected.
    """
    try:
        workbook = load_workbook(workbook_file, data_only=True, read_only=True)
    except Exception as exc:
        raise LevelQuestionImportError(f'Could not read the uploaded file as an .xlsx workbook: {exc}') from exc

    created = []
    failed = []

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue  # empty sheet

        try:
            column_index = _column_index_map(header_row)
        except ValueError as exc:
            failed.append({'sheet': sheet.title, 'row': None, 'reason': str(exc)})
            continue

        for row_number, row in enumerate(rows, start=2):
            if row is None or not any(_cell_text(cell) for cell in row):
                continue  # blank row

            parsed, error = _parse_row(row, column_index)
            if error:
                failed.append({'sheet': sheet.title, 'row': row_number, 'reason': error})
                continue

            _question, question_set = _create_question(assessment_level, parsed)
            created.append({'sheet': sheet.title, 'row': row_number, 'question_set': question_set.label})

    return created, failed
